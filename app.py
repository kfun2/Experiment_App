from flask import Flask, render_template, redirect, url_for, request, jsonify, Response, flash, session, send_file, flash
from forms import NewExperiment, ProductionBatchRecord, GrowthData, ProductionAdditions, AdditionPrcoessVersion, CellBank, SeedProcessCondition, SeedGrowthData
from sqlalchemy import create_engine, bindparam
from sqlalchemy.orm import scoped_session, sessionmaker, Session
from sqlalchemy.sql import text
from sqlalchemy.inspection import inspect
from flask_bootstrap import Bootstrap
import secrets
import csv
import sys
import pandas as pd
import psycopg2
from psycopg2.pool import ThreadedConnectionPool
import psycopg2.extras
from datetime import date
import openpyxl
from openpyxl import load_workbook
import numpy as np
import os
import xlsxwriter
import xlrd
from io import StringIO
from sqlalchemy.sql import text
import io
from flask import make_response


engine = create_engine("postgresql://pd_upstream:hD4RfngcwEf3BK6yyDPS@phoebe:5432/upd_reagent")

sesh = Session(bind=engine)
connection = engine.connect()

#db1 = scoped_session(sessionmaker(bind=engine))

conn = psycopg2.connect(database = "upd_reagent",  user = 'pd_upstream', password = 'hD4RfngcwEf3BK6yyDPS', host = "phoebe", port = "5432", options="-c search_path=dbo,pd_upstream")
db = conn.cursor()

app = Flask(__name__)
Bootstrap(app)

app.config['SECRET_KEY'] = secrets.token_urlsafe(32)
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_TYPE'] = 'filesystem'

@app.route('/')
def home():
    form = NewExperiment()
    return render_template('home.html',form=form)

@app.route('/new_experiment',methods=["Post"])
def new_experiment():
    num_runs = int(request.form.get("num_runs"))
    num_timepoint_lst = []
    for i in range(num_runs):
        num_timepoint_lst.append(int(request.form.get("num_timepoints" + str(i))))
    num_timepoints = sum(num_timepoint_lst)
    phase_lst = ['Pre-Sterilization','Post-Sterilization','Inoculation','Batch Phase','Fed Batch Growth Phase','Fed Batch Induction Phase']
    num_phases = len(phase_lst)

   
    db.execute('SELECT * FROM default_batch_process')
    batchrows = db.fetchall()
    db.execute("Select * FROM default_batch_process LIMIT 0")
    batchkeys = [desc[0] for desc in db.description]
    rows = []
    for batchrow in batchrows:
        rows.append(dict(zip(batchkeys,batchrow)))
    #data = db.execute("SELECT * FROM pd_upstream.default_addition_process")
    #add_rows = [dict(row) for row in data.fetchall()]
    db.execute('SELECT * FROM default_addition_process')
    addrows = db.fetchall()
    db.execute("Select * FROM default_addition_process LIMIT 0")
    addkeys = [desc[0] for desc in db.description]
    add_rows = []
    for addrow in addrows:
        add_rows.append(dict(zip(addkeys,addrow)))


    db.execute("SELECT process_version_name_used_as_template FROM pd_upstream.default_batch_process order by process_version_name_used_as_template")
    default_batch_process_rows = db.fetchall()
    versions = []
    for v in default_batch_process_rows:
        versions.append(v[0])

    db.execute("SELECT process_version_name_used_as_template FROM pd_upstream.default_addition_process order by process_version_name_used_as_template")
    default_add_process_rows = db.fetchall()
    addition_versions = []
    for v in default_add_process_rows:
        if not v[0] in addition_versions:
            addition_versions.append(v[0])

    max_additions = 10
    # keep max_additions at 10 
    return render_template("new_experiment.html",addition_versions=addition_versions,add_rows=add_rows,max_additions=max_additions,num_runs=num_runs,num_timepoint_lst=num_timepoint_lst,num_timepoints=num_timepoints,phase_lst=phase_lst,num_phases=num_phases,rows=rows,versions=versions)

@app.route('/experiment_submitted',methods=['GET', 'POST'])
def experiment_submitted():
 
    dflist = [['23FEB16-1','23FEB16-1', 'Q1' ], ['23FEB16-2','23FEB16-2', 'Q2']]
    insert_sql_df = '''
            INSERT INTO sutro_production_batch_record (process_version_name, run_name, reactor_id)
            VALUES (%s, %s, %s)
            ON CONFLICT (run_name) 
            DO UPDATE SET (process_version_name, run_name, reactor_id) = (Excluded.process_version_name, Excluded.run_name, Excluded.reactor_id)
            '''
    db.executemany(insert_sql_df, dflist)
    
    conn.commit()
    num_runs = int(request.form.get('num_runs'))
    num_phases = int(request.form.get('num_phases'))
    num_timepoint_lst = []

    for run in range(num_runs):
        num_timepoint_lst.append(int(request.form.get('num_timepoint_lst' + str(run))))
    # push experiment data
    experiment_name = request.form.get('experiment_name')
    experiment_name = str(experiment_name)
    today = date.today()
    date_1 = today.strftime("%Y-%m-%d")
    experiment_list = [experiment_name,date_1]
    insert_sql = ''' INSERT INTO pd_upstream.experiment(experiment_name, date1) VALUES (%s,%s) ON CONFLICT (experiment_name) DO UPDATE SET (experiment_name,date1) =(EXCLUDED.experiment_name, EXCLUDED.date1)'''
    

    db.execute(insert_sql,experiment_list)
    conn.commit()
    # push sutro production batch record data
    db.execute("SELECT id FROM pd_upstream.experiment ORDER BY id DESC LIMIT 1")
    experiment_rows =db.fetchall()
    last_id = 0
    for i in experiment_rows:
        last_id = int(i[0])
    
    for i in range(num_runs):
        process_version_name_used_as_template = request.form.get("1-" + str(1+0*num_runs + i))
        process_version_name = request.form.get("1-" + str(1+2*num_runs + i))
        run_name = request.form.get("1-" + str(1+3*num_runs + i))
        reactor_id = request.form.get("1-" + str(1+4*num_runs + i))
        run_description = request.form.get("1-" + str(1+5*num_runs + i))
        process_description = request.form.get("1-" + str(1+6*num_runs + i))
        product = request.form.get("1-" + str(1+7*num_runs + i))
        strain = request.form.get("1-" + str(1+8*num_runs + i))
        site_of_run = request.form.get("1-" + str(1+9*num_runs + i))
        scale_liter = request.form.get("1-" + str(1+10*num_runs + i))
        if scale_liter == "":
            scale_liter = None
        number_of_seed_stages = request.form.get("1-" + str(1+11*num_runs + i))
        if number_of_seed_stages == "":
            number_of_seed_stages = None
        seed_train_run_id = request.form.get("1-" + str(1+12*num_runs + i))
        feed_medium_concentration = request.form.get("1-" + str(1+13*num_runs + i))
        batch_temperature_setpoint_celcius = request.form.get("1-" + str(1+14*num_runs + i))
        if batch_temperature_setpoint_celcius == "":
            batch_temperature_setpoint_celcius = None
        fed_batch_temperature_setpoint_celcius = request.form.get("1-" + str(1+15*num_runs + i))
        if fed_batch_temperature_setpoint_celcius == "":
            fed_batch_temperature_setpoint_celcius = None
        induction_temperature_setpoint_celcius = request.form.get("1-" + str(1+16*num_runs + i))
        if induction_temperature_setpoint_celcius == "":
            induction_temperature_setpoint_celcius = None
        batch_phase_ph_setpoint = request.form.get("1-" + str(1+17*num_runs + i))
        if batch_phase_ph_setpoint == "":
            batch_phase_ph_setpoint = None
        fed_batch_ph_setpoint = request.form.get("1-" + str(1+18*num_runs + i))
        if fed_batch_ph_setpoint == "":
            fed_batch_ph_setpoint = None
        induction_ph_setpoint = request.form.get("1-" + str(1+19*num_runs + i))
        if induction_ph_setpoint == "":
            induction_ph_setpoint = None
        batch_phase_do_setpoint = request.form.get("1-" + str(1+20*num_runs + i))
        if batch_phase_do_setpoint == "":
            batch_phase_do_setpoint = None
        fed_batch_do_setpoint = request.form.get("1-" + str(1+21*num_runs + i))
        if fed_batch_do_setpoint == "":
            fed_batch_do_setpoint = None
        induction_do_setpoint = request.form.get("1-" + str(1+22*num_runs + i))
        if induction_do_setpoint == "":
            induction_do_setpoint = None
        target_batch_phase_airflow_vvm = request.form.get("1-" + str(1+23*num_runs + i))
        if target_batch_phase_airflow_vvm == "":
            target_batch_phase_airflow_vvm = None
        target_fed_batch_phase_airflow_vvm = request.form.get("1-" + str(1+24*num_runs + i))
        if target_fed_batch_phase_airflow_vvm == "":
            target_fed_batch_phase_airflow_vvm = None
        foam_control = request.form.get("1-" + str(1+25*num_runs + i))
        inducer = request.form.get("1-" + str(1+26*num_runs + i))
        target_batch_volume_ml = request.form.get("1-" + str(1+27*num_runs + i))
        if target_batch_volume_ml == "":
            target_batch_volume_ml = None
        target_pre_induction_volume_ml = request.form.get("1-" + str(1+28*num_runs + i))
        if target_pre_induction_volume_ml == "":
            target_pre_induction_volume_ml = None
        target_final_volume_ml = request.form.get("1-" + str(1+29*num_runs + i))
        if target_final_volume_ml == "":
            target_final_volume_ml = None
        target_final_volume_w_drawdown_ml = request.form.get("1-" + str(1+30*num_runs + i))
        if target_final_volume_w_drawdown_ml == "":
            target_final_volume_w_drawdown_ml = None
        seed_media = request.form.get("1-" + str(1+31*num_runs + i))
        feed_media_1 = request.form.get("1-" + str(1+32*num_runs + i))
        feed_media_1_density = request.form.get("1-" + str(1+33*num_runs + i))
        if feed_media_1_density == "":
            feed_media_1_density = None
        target_time_to_add_feed_media_1_post_feed_start_h = request.form.get("1-" + str(1+34*num_runs + i))
        feed_media_2 = request.form.get("1-" + str(1+35*num_runs + i))
        target_time_to_add_feed_media_2_post_feed_start_h = request.form.get("1-" + str(1+36*num_runs + i))
        if target_time_to_add_feed_media_2_post_feed_start_h == "":
            target_time_to_add_feed_media_2_post_feed_start_h = None
        estimated_batch_timing_h = request.form.get("1-" + str(1+37*num_runs + i))
        if estimated_batch_timing_h == "":
            estimated_batch_timing_h = None
        target_post_feed_duration_until_temp_shift_h = request.form.get("1-" + str(1+38*num_runs + i))
        if target_post_feed_duration_until_temp_shift_h == "":
            target_post_feed_duration_until_temp_shift_h = None
        target_post_feed_duration_until_induction_h = request.form.get("1-" + str(1+39*num_runs + i))
        if target_post_feed_duration_until_induction_h == "":
            target_post_feed_duration_until_induction_h = None
        target_induction_duration_h = request.form.get("1-" + str(1+40*num_runs + i))
        if target_induction_duration_h == "":
            target_induction_duration_h = None
        growth_feed_rate_1_m_h_1 = request.form.get("1-" + str(1+41*num_runs + i))
        if growth_feed_rate_1_m_h_1 == "":
            growth_feed_rate_1_m_h_1 = None
        growth_feed_rate_1_post_feed_target_time_h = request.form.get("1-" + str(1+42*num_runs + i))
        if growth_feed_rate_1_post_feed_target_time_h == "":
            growth_feed_rate_1_post_feed_target_time_h = None
        growth_feed_rate_1_type = request.form.get("1-" + str(1+43*num_runs + i))
        growth_feed_rate_2_m_h_1 = request.form.get("1-" + str(1+44*num_runs + i))
        growth_feed_rate_2_post_feed_target_time_h = request.form.get("1-" + str(1+45*num_runs + i))
        growth_feed_rate_2_type = request.form.get("1-" + str(1+46*num_runs + i))
        induction_feed_rate_m_h_1 = request.form.get("1-" + str(1+47*num_runs + i))
        if induction_feed_rate_m_h_1 == "":
            induction_feed_rate_m_h_1 = None
        induction_feed_profile_type = request.form.get("1-" + str(1+48*num_runs + i))
        induction_phase_feed_interval_h = request.form.get("1-" + str(1+49*num_runs + i))
        if induction_phase_feed_interval_h == "":
            induction_phase_feed_interval_h = None
        post_feed_time_of_percent_drawdown_h = request.form.get("1-" + str(1+50*num_runs + i))
        if post_feed_time_of_percent_drawdown_h == "":
            post_feed_time_of_percent_drawdown_h = None
        drawdown_at_post_feed_time_percent = request.form.get("1-" + str(1+51*num_runs + i))
        if drawdown_at_post_feed_time_percent == "":
            drawdown_at_post_feed_time_percent = None
        initial_batch_vol_litter = request.form.get("1-" + str(1+52*num_runs + i))
        if initial_batch_vol_litter == "":
            initial_batch_vol_litter = None
        x0_grams_over_litter = request.form.get("1-" + str(1+53*num_runs + i))
        if x0_grams_over_litter == "":
            x0_grams_over_litter = None
        sf_feed_gluc_grams_over_litter = request.form.get("1-" + str(1+54*num_runs + i))
        if sf_feed_gluc_grams_over_litter == "":
            sf_feed_gluc_grams_over_litter = None
        s0_batch_gluc_grams_over_litter = request.form.get("1-" + str(1+55*num_runs + i))
        if s0_batch_gluc_grams_over_litter == "":
            s0_batch_gluc_grams_over_litter = None
        yx_over_s_grams_over_grams = request.form.get("1-" + str(1+56*num_runs + i))
        if yx_over_s_grams_over_grams == "":
            yx_over_s_grams_over_grams = None
        od_over_gram_dry_wt_over_litter = request.form.get("1-" + str(1+57*num_runs + i))    
        if od_over_gram_dry_wt_over_litter == "":
            od_over_gram_dry_wt_over_litter = None
        growth_phase_feed_interval_h = request.form.get("1-" + str(1+58*num_runs + i))
        if growth_phase_feed_interval_h == "":
            growth_phase_feed_interval_h = None
        feed_profile_type = request.form.get("1-" + str(1+59*num_runs + i))
        actual_inoculation_time = request.form.get("1-" + str(1+60*num_runs + i))
        if actual_inoculation_time == "":
            actual_inoculation_time = None
        actual_fed_batch_start_time = request.form.get("1-" + str(1+61*num_runs + i))
        if actual_fed_batch_start_time == "":
            actual_fed_batch_start_time = None
        actual_induction_start_time = request.form.get("1-" + str(1+62*num_runs + i))
        if actual_induction_start_time == "":
            actual_induction_start_time = None
        target_induction_duration_1_h = request.form.get("1-" + str(1+63*num_runs + i))
        if target_induction_duration_1_h == "":
            target_induction_duration_1_h = None
        addition_on_scale_pump_1 = request.form.get("1-" + str(1+64*num_runs + i))
        if addition_on_scale_pump_1 == "":
            addition_on_scale_pump_1 = None
        addition_on_scale_pump_2 = request.form.get("1-" + str(1+65*num_runs + i))
        if addition_on_scale_pump_2 == "":
            addition_on_scale_pump_2 = None
        run_outcome = request.form.get("1-" + str(1+66*num_runs + i))
        db.execute("INSERT INTO pd_upstream.sutro_production_batch_record (process_description,process_version_name,target_batch_volume_ml,process_version_name_used_as_template,run_name,reactor_id,run_description,product,strain,site_of_run,scale_liter,number_of_seed_stages,seed_train_run_id,\
                    feed_medium_concentration,batch_temperature_setpoint_celcius,fed_batch_temperature_setpoint_celcius,induction_temperature_setpoint_celcius,\
                    batch_phase_ph_setpoint,fed_batch_ph_setpoint,induction_ph_setpoint,batch_phase_do_setpoint,fed_batch_do_setpoint,induction_do_setpoint,target_batch_phase_airflow_vvm,target_fed_batch_phase_airflow_vvm,foam_control,inducer,\
                    target_pre_induction_volume_ml,target_final_volume_ml,target_final_volume_w_drawdown_ml,seed_media,feed_media_1,feed_media_1_density,\
                    target_time_to_add_feed_media_1_post_feed_start_h,feed_media_2,target_time_to_add_feed_media_2_post_feed_start_h,estimated_batch_timing_h,\
                    target_post_feed_duration_until_temp_shift_h,target_post_feed_duration_until_induction_h,target_induction_duration_h,growth_feed_rate_1_m_h_1,\
                    growth_feed_rate_1_post_feed_target_time_h,growth_feed_rate_1_type,growth_feed_rate_2_m_h_1,growth_feed_rate_2_post_feed_target_time_h,\
                    growth_feed_rate_2_type,induction_feed_rate_m_h_1,induction_feed_profile_type,post_feed_time_of_percent_drawdown_h,drawdown_at_post_feed_time_percent,\
                    initial_batch_vol_litter,x0_grams_over_litter,sf_feed_gluc_grams_over_litter,s0_batch_gluc_grams_over_litter,yx_over_s_grams_over_grams,\
                    od_over_gram_dry_wt_over_litter,growth_phase_feed_interval_h,feed_profile_type,induction_phase_feed_interval_h,actual_inoculation_time,\
                    actual_fed_batch_start_time,actual_induction_start_time,target_induction_duration_1_h,run_outcome,experiment_id,addition_on_scale_pump_1,addition_on_scale_pump_2)\
                    VALUES (%(process_description)s, %(process_version_name)s, %(target_batch_volume_ml)s, %(process_version_name_used_as_template)s, %(run_name)s, %(reactor_id)s, %(run_description)s, %(product)s, %(strain)s, %(site_of_run)s, %(scale_liter)s, %(number_of_seed_stages)s, %(seed_train_run_id)s, %(feed_medium_concentration)s, %(batch_temperature_setpoint_celcius)s, %(fed_batch_temperature_setpoint_celcius)s, %(induction_temperature_setpoint_celcius)s, %(batch_phase_ph_setpoint)s, %(fed_batch_ph_setpoint)s, %(induction_ph_setpoint)s, %(batch_phase_do_setpoint)s, %(fed_batch_do_setpoint)s, %(induction_do_setpoint)s, %(target_batch_phase_airflow_vvm)s, %(target_fed_batch_phase_airflow_vvm)s, %(foam_control)s, %(inducer)s, %(target_pre_induction_volume_ml)s, %(target_final_volume_ml)s, %(target_final_volume_w_drawdown_ml)s, %(seed_media)s, %(feed_media_1)s, %(feed_media_1_density)s, %(target_time_to_add_feed_media_1_post_feed_start_h)s, %(feed_media_2)s, %(target_time_to_add_feed_media_2_post_feed_start_h)s, %(estimated_batch_timing_h)s, %(target_post_feed_duration_until_temp_shift_h)s, %(target_post_feed_duration_until_induction_h)s, %(target_induction_duration_h)s, %(growth_feed_rate_1_m_h_1)s, %(growth_feed_rate_1_post_feed_target_time_h)s, %(growth_feed_rate_1_type)s, %(growth_feed_rate_2_m_h_1)s, %(growth_feed_rate_2_post_feed_target_time_h)s, %(growth_feed_rate_2_type)s, %(induction_feed_rate_m_h_1)s, %(induction_feed_profile_type)s, %(post_feed_time_of_percent_drawdown_h)s, %(drawdown_at_post_feed_time_percent)s, %(initial_batch_vol_litter)s, %(x0_grams_over_litter)s, %(sf_feed_gluc_grams_over_litter)s, %(s0_batch_gluc_grams_over_litter)s, %(yx_over_s_grams_over_grams)s, %(od_over_gram_dry_wt_over_litter)s, %(growth_phase_feed_interval_h)s, %(feed_profile_type)s, %(induction_phase_feed_interval_h)s, %(actual_inoculation_time)s, %(actual_fed_batch_start_time)s, %(actual_induction_start_time)s, %(target_induction_duration_1_h)s, %(run_outcome)s, %(experiment_id)s, %(addition_on_scale_pump_1)s, %(addition_on_scale_pump_2)s)\
                    On conflict (run_name) DO UPDATE SET \
                    (process_description,process_version_name,target_batch_volume_ml,process_version_name_used_as_template,run_name,reactor_id,run_description,product,strain,site_of_run,scale_liter,number_of_seed_stages,seed_train_run_id,\
                    feed_medium_concentration,batch_temperature_setpoint_celcius,fed_batch_temperature_setpoint_celcius,induction_temperature_setpoint_celcius,\
                    batch_phase_ph_setpoint,fed_batch_ph_setpoint,induction_ph_setpoint,batch_phase_do_setpoint,fed_batch_do_setpoint,induction_do_setpoint,target_batch_phase_airflow_vvm,target_fed_batch_phase_airflow_vvm,foam_control,inducer,\
                    target_pre_induction_volume_ml,target_final_volume_ml,target_final_volume_w_drawdown_ml,seed_media,feed_media_1,feed_media_1_density,\
                    target_time_to_add_feed_media_1_post_feed_start_h,feed_media_2,target_time_to_add_feed_media_2_post_feed_start_h,estimated_batch_timing_h,\
                    target_post_feed_duration_until_temp_shift_h,target_post_feed_duration_until_induction_h,target_induction_duration_h,growth_feed_rate_1_m_h_1,\
                    growth_feed_rate_1_post_feed_target_time_h,growth_feed_rate_1_type,growth_feed_rate_2_m_h_1,growth_feed_rate_2_post_feed_target_time_h,\
                    growth_feed_rate_2_type,induction_feed_rate_m_h_1,induction_feed_profile_type,post_feed_time_of_percent_drawdown_h,drawdown_at_post_feed_time_percent,\
                    initial_batch_vol_litter,x0_grams_over_litter,sf_feed_gluc_grams_over_litter,s0_batch_gluc_grams_over_litter,yx_over_s_grams_over_grams,\
                    od_over_gram_dry_wt_over_litter,growth_phase_feed_interval_h,feed_profile_type,induction_phase_feed_interval_h,actual_inoculation_time,\
                    actual_fed_batch_start_time,actual_induction_start_time,target_induction_duration_1_h,run_outcome,experiment_id,addition_on_scale_pump_1,addition_on_scale_pump_2) = \
                    (EXCLUDED.process_description, EXCLUDED.process_version_name, EXCLUDED.target_batch_volume_ml, EXCLUDED.process_version_name_used_as_template, EXCLUDED.run_name, EXCLUDED.reactor_id, EXCLUDED.run_description, EXCLUDED.product, EXCLUDED.strain, EXCLUDED.site_of_run, EXCLUDED.scale_liter, EXCLUDED.number_of_seed_stages, EXCLUDED.seed_train_run_id, EXCLUDED. feed_medium_concentration, EXCLUDED.batch_temperature_setpoint_celcius, EXCLUDED.fed_batch_temperature_setpoint_celcius, EXCLUDED.induction_temperature_setpoint_celcius, EXCLUDED. batch_phase_ph_setpoint, EXCLUDED.fed_batch_ph_setpoint, EXCLUDED.induction_ph_setpoint, EXCLUDED.batch_phase_do_setpoint, EXCLUDED.fed_batch_do_setpoint, EXCLUDED.induction_do_setpoint, EXCLUDED.target_batch_phase_airflow_vvm, EXCLUDED.target_fed_batch_phase_airflow_vvm, EXCLUDED.foam_control, EXCLUDED.inducer, EXCLUDED. target_pre_induction_volume_ml, EXCLUDED.target_final_volume_ml, EXCLUDED.target_final_volume_w_drawdown_ml, EXCLUDED.seed_media, EXCLUDED.feed_media_1, EXCLUDED.feed_media_1_density, EXCLUDED. target_time_to_add_feed_media_1_post_feed_start_h, EXCLUDED.feed_media_2, EXCLUDED.target_time_to_add_feed_media_2_post_feed_start_h, EXCLUDED.estimated_batch_timing_h, EXCLUDED. target_post_feed_duration_until_temp_shift_h, EXCLUDED.target_post_feed_duration_until_induction_h, EXCLUDED.target_induction_duration_h, EXCLUDED.growth_feed_rate_1_m_h_1, EXCLUDED. growth_feed_rate_1_post_feed_target_time_h, EXCLUDED.growth_feed_rate_1_type, EXCLUDED.growth_feed_rate_2_m_h_1, EXCLUDED.growth_feed_rate_2_post_feed_target_time_h, EXCLUDED.growth_feed_rate_2_type, EXCLUDED.induction_feed_rate_m_h_1, EXCLUDED.induction_feed_profile_type, EXCLUDED.post_feed_time_of_percent_drawdown_h, EXCLUDED.drawdown_at_post_feed_time_percent, EXCLUDED. initial_batch_vol_litter, EXCLUDED.x0_grams_over_litter, EXCLUDED.sf_feed_gluc_grams_over_litter, EXCLUDED.s0_batch_gluc_grams_over_litter, EXCLUDED.yx_over_s_grams_over_grams, EXCLUDED. od_over_gram_dry_wt_over_litter, EXCLUDED.growth_phase_feed_interval_h, EXCLUDED.feed_profile_type, EXCLUDED.induction_phase_feed_interval_h, EXCLUDED.actual_inoculation_time, EXCLUDED.actual_fed_batch_start_time, EXCLUDED.actual_induction_start_time, EXCLUDED.target_induction_duration_1_h, EXCLUDED.run_outcome, EXCLUDED.experiment_id, EXCLUDED.addition_on_scale_pump_1, EXCLUDED.addition_on_scale_pump_2) ",
                {"process_description":process_description,"process_version_name":process_version_name,"target_batch_volume_ml":target_batch_volume_ml,"addition_on_scale_pump_1":addition_on_scale_pump_1,"addition_on_scale_pump_2":addition_on_scale_pump_2,"feed_media_1_density":feed_media_1_density,"process_version_name_used_as_template":process_version_name_used_as_template,"run_name":run_name,"reactor_id":reactor_id,"run_description":run_description,"product":product,"strain":strain,"site_of_run":site_of_run,"scale_liter":scale_liter,"number_of_seed_stages":number_of_seed_stages,"seed_train_run_id":seed_train_run_id,"feed_medium_concentration":feed_medium_concentration,"batch_temperature_setpoint_celcius":batch_temperature_setpoint_celcius,"fed_batch_temperature_setpoint_celcius":fed_batch_temperature_setpoint_celcius,"induction_temperature_setpoint_celcius":induction_temperature_setpoint_celcius,"batch_phase_ph_setpoint":batch_phase_ph_setpoint,"fed_batch_ph_setpoint":fed_batch_ph_setpoint,"induction_ph_setpoint":induction_ph_setpoint,"batch_phase_do_setpoint":batch_phase_do_setpoint,"fed_batch_do_setpoint":fed_batch_do_setpoint,"induction_do_setpoint":induction_do_setpoint,"target_batch_phase_airflow_vvm":target_batch_phase_airflow_vvm,"target_fed_batch_phase_airflow_vvm":target_fed_batch_phase_airflow_vvm,"foam_control":foam_control,"inducer":inducer,"target_pre_induction_volume_ml":target_pre_induction_volume_ml,"target_final_volume_ml":target_final_volume_ml,"target_final_volume_w_drawdown_ml":target_final_volume_w_drawdown_ml,"seed_media":seed_media,"feed_media_1":feed_media_1,"target_time_to_add_feed_media_1_post_feed_start_h":target_time_to_add_feed_media_1_post_feed_start_h,"feed_media_2":feed_media_2,"target_time_to_add_feed_media_2_post_feed_start_h":target_time_to_add_feed_media_2_post_feed_start_h,"estimated_batch_timing_h":estimated_batch_timing_h,"target_post_feed_duration_until_temp_shift_h":target_post_feed_duration_until_temp_shift_h,"target_post_feed_duration_until_induction_h":target_post_feed_duration_until_induction_h,"target_induction_duration_h":target_induction_duration_h,"growth_feed_rate_1_m_h_1":growth_feed_rate_1_m_h_1,"growth_feed_rate_1_post_feed_target_time_h":growth_feed_rate_1_post_feed_target_time_h,"growth_feed_rate_1_type":growth_feed_rate_1_type,"growth_feed_rate_2_m_h_1":growth_feed_rate_2_m_h_1,"growth_feed_rate_2_post_feed_target_time_h":growth_feed_rate_2_post_feed_target_time_h,"growth_feed_rate_2_type":growth_feed_rate_2_type,"induction_feed_rate_m_h_1":induction_feed_rate_m_h_1,"induction_feed_profile_type":induction_feed_profile_type,"post_feed_time_of_percent_drawdown_h":post_feed_time_of_percent_drawdown_h,"drawdown_at_post_feed_time_percent":drawdown_at_post_feed_time_percent,"initial_batch_vol_litter":initial_batch_vol_litter,"x0_grams_over_litter":x0_grams_over_litter,"sf_feed_gluc_grams_over_litter":sf_feed_gluc_grams_over_litter,"s0_batch_gluc_grams_over_litter":s0_batch_gluc_grams_over_litter,"yx_over_s_grams_over_grams":yx_over_s_grams_over_grams,"od_over_gram_dry_wt_over_litter":od_over_gram_dry_wt_over_litter,"growth_phase_feed_interval_h":growth_phase_feed_interval_h,"feed_profile_type":feed_profile_type,"induction_phase_feed_interval_h":induction_phase_feed_interval_h,"actual_inoculation_time":actual_inoculation_time,"actual_fed_batch_start_time":actual_fed_batch_start_time,"actual_induction_start_time":actual_induction_start_time,"target_induction_duration_1_h":target_induction_duration_1_h,"run_outcome":run_outcome,"experiment_id":last_id}) 
        conn.commit()
    # push production additions data
    on_off_lst = []
    for r in range(num_phases):
        on_off_lst.append(request.form.get('phase_on_or_off' + str(r)))
    num_additions = int(request.form.get('num_additions'))
    db.execute("SELECT id FROM pd_upstream.sutro_production_batch_record ORDER BY id DESC LIMIT {}".format(num_runs))
    last_id_q = db.fetchall()
    last_ids = []
    for i in last_id_q:
        last_ids.append(int(i[0]))
    last_ids.reverse()
    run_name_lst = []
    process_version_name_lst = []
    process_version_name_actual_lst = []
    num_additions_by_phase = []
    initial_batch_vol_ml_lst = []
    water_gain_loss_lst = []
    for k in range(num_runs):
        run_name_lst.append(request.form.get("2-" + str(k + 1)))
        process_version_name_lst.append(request.form.get("2-" + str(k + (num_runs + 1))))
        process_version_name_actual_lst.append(request.form.get("2-" + str(k + (3*num_runs + 1))))
        initial_batch_vol_ml_lst.append(request.form.get("2-" + str(k + (4*num_runs + 1))))
        water_gain_loss_lst.append(request.form.get("2-" + str(k + (5*num_runs + 1))))
    for l in range(num_phases):
        num_additions_by_phase.append(int(request.form.get('num_additions_lst' + str(l))))
    starting_index = 6*num_runs + 1
    for q in range(num_runs):
        num_seen = 0
        for i in range(10*num_phases):
            if on_off_lst[i//10] == 'off':
                if (i+1) % 10 == 0:
                    num_seen += 1
                continue
            else:
                if ((i + 1) - 10*num_seen) > num_additions_by_phase[i//10]:
                    if (i+1) % 10 == 0:
                        num_seen += 1
                    continue
                run_id = run_name_lst[q]
                process_version_name_used_as_template = process_version_name_lst[q]
                process_version_name = process_version_name_actual_lst[q]
                initial_batch_vol_ml = initial_batch_vol_ml_lst[q]
                if initial_batch_vol_ml == "":
                    initial_batch_vol_ml = None
                water_gain_loss = water_gain_loss_lst[q]
                if water_gain_loss == "":
                    water_gain_loss = None
                addition_name = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*0 + q))
                target_phase_of_addition = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*1 + q))
                nth_addition = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*2 + q))
                if nth_addition == "":
                    nth_addition = None
                target_batch_concentration_ml_li = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*3 + q))
                if target_batch_concentration_ml_li == "":
                    target_batch_concentration_ml_li = None
                target_amount_ml = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*4 + q))
                if target_amount_ml == "":
                    target_amount_ml = None
                target_phase_time_h = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*5 + q))
                if target_phase_time_h == "":
                    target_phase_time_h = None
                actual_amount_added_ml = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*6 + q))
                if actual_amount_added_ml == "":
                    actual_amount_added_ml = None
                actual_start_time_of_addition = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*7 + q))
                if actual_start_time_of_addition == "":
                    actual_start_time_of_addition = None
                actual_stop_time_of_addition = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*8 + q))
                if actual_stop_time_of_addition == "":
                    actual_stop_time_of_addition = None
                addition_lot_id = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*9 + q))
                addition_type = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*10 + q))
                target_addition_rate_ml_h_li = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*11 + q))
                if target_addition_rate_ml_h_li == "":
                    target_addition_rate_ml_h_li = None
                target_addition_rate_ml_h = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*12 + q))
                if target_addition_rate_ml_h == "":
                    target_addition_rate_ml_h = None
                actual_addition_rate_ml_h = request.form.get("2-" + str(starting_index + i*14*num_runs + num_runs*13 + q))
                if actual_addition_rate_ml_h == "":
                    actual_addition_rate_ml_h = None
                db.execute("INSERT INTO pd_upstream.production_addition (process_version_name,initial_batch_vol_ml,water_gain_loss,run_id,process_version_name_used_as_template,target_phase_of_addition,addition_name, \
                            nth_addition,target_batch_concentration_ml_li,target_amount_ml,target_phase_time_h, \
                            actual_amount_added_ml,actual_start_time_of_addition,actual_stop_time_of_addition, \
                            addition_lot_id,addition_type,target_addition_rate_ml_h_li,target_addition_rate_ml_h,actual_addition_rate_ml_h,actual_run_id) \
                            VALUES (:process_version_name,:initial_batch_vol_ml,:water_gain_loss,:run_id,:process_version_name_used_as_template,:target_phase_of_addition,:addition_name, \
                            :nth_addition,:target_batch_concentration_ml_li,:target_amount_ml,:target_phase_time_h, \
                            :actual_amount_added_ml,:actual_start_time_of_addition,:actual_stop_time_of_addition, \
                            :addition_lot_id,:addition_type,:target_addition_rate_ml_h_li,:target_addition_rate_ml_h,:actual_addition_rate_ml_h,:last_id)",
                    {"process_version_name":process_version_name,"initial_batch_vol_ml":initial_batch_vol_ml,"water_gain_loss":water_gain_loss,"actual_addition_rate_ml_h":actual_addition_rate_ml_h,"target_phase_of_addition":target_phase_of_addition,"run_id":run_id,"process_version_name_used_as_template":process_version_name_used_as_template,"addition_name":addition_name,"nth_addition":nth_addition,"target_batch_concentration_ml_li":target_batch_concentration_ml_li,"target_amount_ml":target_amount_ml,"target_phase_time_h":target_phase_time_h,"actual_amount_added_ml":actual_amount_added_ml,"actual_start_time_of_addition":actual_start_time_of_addition,"actual_stop_time_of_addition":actual_stop_time_of_addition,"addition_lot_id":addition_lot_id,"addition_type":addition_type,"target_addition_rate_ml_h_li":target_addition_rate_ml_h_li,"target_addition_rate_ml_h":target_addition_rate_ml_h,"last_id":last_ids[q]}) 
                conn.commit()
                if (i+1) % 10 == 0:
                    num_seen += 1
            if (i+1) % 10 == 0:
                    num_seen += 1
    # push growth data
    db.execute("SELECT id FROM pd_upstream.sutro_production_batch_record ORDER BY id DESC LIMIT {}".format(num_runs))
    last_id_q = db.fetchall()
    last_ids = []
    counter = 0
    for i in last_id_q:
        last_ids.append(int(i[0]))
    last_ids.reverse()
    for i in range(num_runs):
        prev = 0
        if i > 0:
            prev = num_timepoint_lst[i-1]
        for j in range(int(num_timepoint_lst[i])):
            run_id = request.form.get("3-" + str(1 +  + counter*54))
            ferm_sample_id = request.form.get("3-" + str(2 + counter*54))
            reactor_id_temp = request.form.get("3-" + str(3 + counter*54))
            ferm_stage = request.form.get("3-" + str(4 + counter*54))
            datetime = request.form.get("3-" + str(5 + counter*54))
            if datetime == "":
                datetime = None
            notes = request.form.get("3-" + str(6 + counter*54))
            eft_h = request.form.get("3-" + str(7 + counter*54))
            if eft_h == "":
                eft_h = None
            time_post_induction_h = request.form.get("3-" + str(8 + counter*54))
            if time_post_induction_h == "":
                time_post_induction_h = None
            rounded_post_induction_time_h = request.form.get("3-" + str(9 + counter*54))
            if rounded_post_induction_time_h == "":
                rounded_post_induction_time_h = None
            post_feed_start_time_h = request.form.get("3-" + str(10 + counter*54))
            if post_feed_start_time_h == "":
                post_feed_start_time_h = None
            o2_flow_percent = request.form.get("3-" + str(11 + counter*54))
            if o2_flow_percent == "":
                o2_flow_percent = None
            total_gas_flow_slpm = request.form.get("3-" + str(12 + counter*54))
            if total_gas_flow_slpm == "":
                total_gas_flow_slpm = None
            base_totalizer_ml = request.form.get("3-" + str(13 + counter*54))
            if base_totalizer_ml == "":
                base_totalizer_ml = None
            base_consumed_g = request.form.get("3-" + str(14 + counter*54))
            if base_consumed_g == "":
                base_consumed_g = None
            feed_weight_g = request.form.get("3-" + str(15 + counter*54))
            if feed_weight_g == "":
                feed_weight_g = None
            amount_of_feed_added_g = request.form.get("3-" + str(16 + counter*54))
            if amount_of_feed_added_g == "":
                amount_of_feed_added_g = None
            amount_of_feed_added_ml = request.form.get("3-" + str(17 + counter*54))
            if amount_of_feed_added_ml == "":
                amount_of_feed_added_ml = None
            antifoam_weight_g = request.form.get("3-" + str(18 + counter*54))
            if antifoam_weight_g == "":
                antifoam_weight_g = None
            scale_1_weight_g = request.form.get("3-" + str(19 + counter*54))
            if scale_1_weight_g == "":
                scale_1_weight_g = None
            pump_1_rpm = request.form.get("3-" + str(20 + counter*54))
            if pump_1_rpm == "":
                pump_1_rpm = None
            rate_from_scale_1_g_h = request.form.get("3-" + str(21 + counter*54))
            if rate_from_scale_1_g_h == "":
                rate_from_scale_1_g_h = None
            scale_2_weight_g = request.form.get("3-" + str(22 + counter*54))
            if scale_2_weight_g == "":
                scale_2_weight_g = None
            pump_2_rpm = request.form.get("3-" + str(23 + counter*54))
            if pump_2_rpm == "":
                pump_2_rpm = None
            rate_from_scale_2_g_h = request.form.get("3-" + str(24 + counter*54))
            if rate_from_scale_2_g_h == "":
                rate_from_scale_2_g_h = None
            offline_ph = request.form.get("3-" + str(25 + counter*54))
            if offline_ph == "":
                offline_ph = None
            od595 = request.form.get("3-" + str(26 + counter*54))
            if od595 == "":
                od595 = None
            od_bioht = request.form.get("3-" + str(27 + counter*54))
            if od_bioht == "":
                od_bioht = None
            bioht_glucose_g_l = request.form.get("3-" + str(28 + counter*54))
            if bioht_glucose_g_l == "":
                bioht_glucose_g_l = None
            bioht_lactate_mmol = request.form.get("3-" + str(29 + counter*54))
            if bioht_lactate_mmol == "":
                bioht_lactate_mmol = None
            bioht_acetate_mmol = request.form.get("3-" + str(30 + counter*54))
            if bioht_acetate_mmol == "":
                bioht_acetate_mmol = None
            bioht_ammonia_mmol = request.form.get("3-" + str(31 + counter*54))
            if bioht_ammonia_mmol == "":
                bioht_ammonia_mmol = None
            bioht_glutamine_mmol = request.form.get("3-" + str(32 + counter*54))
            if bioht_glutamine_mmol == "":
                bioht_glutamine_mmol = None
            bioht_glutamate_mmol = request.form.get("3-" + str(33 + counter*54))
            if bioht_glutamate_mmol == "":
                bioht_glutamate_mmol = None
            bioht_phosphate_mmol = request.form.get("3-" + str(34 + counter*54))
            if bioht_phosphate_mmol == "":
                bioht_phosphate_mmol = None
            bioht_magnesium_mmol = request.form.get("3-" + str(35 + counter*54))
            if bioht_magnesium_mmol == "":
                bioht_magnesium_mmol = None
            osm_mosm_kg = request.form.get("3-" + str(36 + counter*54))
            if osm_mosm_kg == "":
                osm_mosm_kg = None
            bioht_arabinose_mg_l = request.form.get("3-" + str(37 + counter*54))
            if bioht_arabinose_mg_l == "":
                bioht_arabinose_mg_l = None
            af_added_ml = request.form.get("3-" + str(38 + counter*54))
            if af_added_ml == "":
                af_added_ml = None
            af_totalizer_ml = request.form.get("3-" + str(39 + counter*54))
            if af_totalizer_ml == "":
                af_totalizer_ml = None
            theoretical_amount_of_feed_added_ml = request.form.get("3-" + str(40 + counter*54))
            if theoretical_amount_of_feed_added_ml == "":
                theoretical_amount_of_feed_added_ml = None
            percent_diff_in_feed = request.form.get("3-" + str(41 + counter*54))
            if percent_diff_in_feed == "":
                percent_diff_in_feed = None
            online_ph = request.form.get("3-" + str(42 + counter*54))
            if online_ph == "":
                online_ph = None
            broth_viscosity_ranking = request.form.get("3-" + str(43 + counter*54))
            if broth_viscosity_ranking == "":
                broth_viscosity_ranking = None
            bioht_pyruvate_mmol = request.form.get("3-" + str(44 + counter*54))
            if bioht_pyruvate_mmol == "":
                bioht_pyruvate_mmol = None
            total_amount_of_antifoam_added_ml = request.form.get("3-" + str(45 + counter*54))
            if total_amount_of_antifoam_added_ml == "":
                total_amount_of_antifoam_added_ml = None
            bioht_ldh_u_l = request.form.get("3-" + str(46 + counter*54))
            if bioht_ldh_u_l == "":
                bioht_ldh_u_l = None
            bioht_igg_mg_l = request.form.get("3-" + str(47 + counter*54))
            if bioht_igg_mg_l == "":
                bioht_igg_mg_l = None
            bioht_total_protein_g_l = request.form.get("3-" + str(48 + counter*54))
            if bioht_total_protein_g_l == "":
                bioht_total_protein_g_l = None
            sodium = request.form.get("3-" + str(49 + counter*54))
            if sodium == "":
                sodium = None
            potassium = request.form.get("3-" + str(50 + counter*54))
            if potassium == "":
                potassium = None
            bioht_formate_mg_l = request.form.get("3-" + str(51 + counter*54))
            if bioht_formate_mg_l == "":
                bioht_formate_mg_l = None
            percent_co2 = request.form.get("3-" + str(52 + counter*54))
            if percent_co2 == "":
                percent_co2 = None
            bioht_glycerol_mg_l = request.form.get("3-" + str(53 + counter*54))
            if bioht_glycerol_mg_l == "":
                bioht_glycerol_mg_l = None
            gel_titer_g_l = request.form.get("3-" + str(54 + counter*54))
            if gel_titer_g_l == "":
                gel_titer_g_l = None
            counter +=1
            db.execute("INSERT INTO pd_upstream.growth_data \
                        (run_id,ferm_sample_id,reactor_id_temp,ferm_stage,notes,eft_h,time_post_induction_h,rounded_post_induction_time_h, \
                        post_feed_start_time_h,o2_flow_percent,total_gas_flow_slpm,base_totalizer_ml,base_consumed_g,feed_weight_g,amount_of_feed_added_g, \
                        amount_of_feed_added_ml,offline_ph,od595,od_bioht,bioht_glucose_g_l,bioht_lactate_mmol,bioht_acetate_mmol,bioht_ammonia_mmol, \
                        bioht_glutamine_mmol,bioht_glutamate_mmol,bioht_phosphate_mmol,bioht_magnesium_mmol,osm_mosm_kg,bioht_arabinose_mg_l,af_added_ml, \
                        af_totalizer_ml,theoretical_amount_of_feed_added_ml,percent_diff_in_feed,online_ph,broth_viscosity_ranking,bioht_pyruvate_mmol, \
                        total_amount_of_antifoam_added_ml,bioht_ldh_u_l,bioht_igg_mg_l,bioht_total_protein_g_l,sodium,potassium,bioht_formate_mg_l, \
                        percent_co2,bioht_glycerol_mg_l,datetime,gel_titer_g_l,antifoam_weight_g,scale_1_weight_g,pump_1_rpm,rate_from_scale_1_g_h,scale_2_weight_g,pump_2_rpm,rate_from_scale_2_g_h) VALUES (%(run_id)s, %(ferm_sample_id)s, %(reactor_id_temp)s, %(ferm_stage)s, %(notes)s, %(eft_h)s, %(time_post_induction_h)s, %(rounded_post_induction_time_h)s, %(post_feed_start_time_h)s, %(o2_flow_percent)s, %(total_gas_flow_slpm)s, %(base_totalizer_ml)s, %(base_consumed_g)s, %(feed_weight_g)s, %(amount_of_feed_added_g)s, %(amount_of_feed_added_ml)s, %(offline_ph)s, %(od595)s, %(od_bioht)s, %(bioht_glucose_g_l)s, %(bioht_lactate_mmol)s, %(bioht_acetate_mmol)s, %(bioht_ammonia_mmol)s, %(bioht_glutamine_mmol)s, %(bioht_glutamate_mmol)s, %(bioht_phosphate_mmol)s, %(bioht_magnesium_mmol)s, %(osm_mosm_kg)s, %(bioht_arabinose_mg_l)s, %(af_added_ml)s, %(af_totalizer_ml)s, %(theoretical_amount_of_feed_added_ml)s, %(percent_diff_in_feed)s, %(online_ph)s, %(broth_viscosity_ranking)s, %(bioht_pyruvate_mmol)s, %(total_amount_of_antifoam_added_ml)s, %(bioht_ldh_u_l)s, %(bioht_igg_mg_l)s, %(bioht_total_protein_g_l)s, %(sodium)s, %(potassium)s, %(bioht_formate_mg_l)s, %(percent_co2)s, %(bioht_glycerol_mg_l)s, %(datetime)s, %(gel_titer_g_l)s, %(antifoam_weight_g)s, %(scale_1_weight_g)s, %(pump_1_rpm)s, %(rate_from_scale_1_g_h)s, %(scale_2_weight_g)s, %(pump_2_rpm)s, %(rate_from_scale_2_g_h)s)\
                        ON CONFLICT (run_id,ferm_sample_id) DO UPDATE SET\
                        (run_id,ferm_sample_id,reactor_id_temp,ferm_stage,notes,eft_h,time_post_induction_h,rounded_post_induction_time_h, \
                        post_feed_start_time_h,o2_flow_percent,total_gas_flow_slpm,base_totalizer_ml,base_consumed_g,feed_weight_g,amount_of_feed_added_g, \
                        amount_of_feed_added_ml,offline_ph,od595,od_bioht,bioht_glucose_g_l,bioht_lactate_mmol,bioht_acetate_mmol,bioht_ammonia_mmol, \
                        bioht_glutamine_mmol,bioht_glutamate_mmol,bioht_phosphate_mmol,bioht_magnesium_mmol,osm_mosm_kg,bioht_arabinose_mg_l,af_added_ml, \
                        af_totalizer_ml,theoretical_amount_of_feed_added_ml,percent_diff_in_feed,online_ph,broth_viscosity_ranking,bioht_pyruvate_mmol, \
                        total_amount_of_antifoam_added_ml,bioht_ldh_u_l,bioht_igg_mg_l,bioht_total_protein_g_l,sodium,potassium,bioht_formate_mg_l, \
                        percent_co2,bioht_glycerol_mg_l,datetime,gel_titer_g_l,antifoam_weight_g,scale_1_weight_g,pump_1_rpm,rate_from_scale_1_g_h,scale_2_weight_g,pump_2_rpm,rate_from_scale_2_g_h) = \
                        (EXCLUDED.run_id, EXCLUDED.ferm_sample_id, EXCLUDED.reactor_id_temp, EXCLUDED.ferm_stage, EXCLUDED.notes, EXCLUDED.eft_h, EXCLUDED.time_post_induction_h, EXCLUDED.rounded_post_induction_time_h, EXCLUDED. post_feed_start_time_h, EXCLUDED.o2_flow_percent, EXCLUDED.total_gas_flow_slpm, EXCLUDED.base_totalizer_ml, EXCLUDED.base_consumed_g, EXCLUDED.feed_weight_g, EXCLUDED.amount_of_feed_added_g, EXCLUDED. amount_of_feed_added_ml, EXCLUDED.offline_ph, EXCLUDED.od595, EXCLUDED.od_bioht, EXCLUDED.bioht_glucose_g_l, EXCLUDED.bioht_lactate_mmol, EXCLUDED.bioht_acetate_mmol, EXCLUDED.bioht_ammonia_mmol, EXCLUDED. bioht_glutamine_mmol, EXCLUDED.bioht_glutamate_mmol, EXCLUDED.bioht_phosphate_mmol, EXCLUDED.bioht_magnesium_mmol, EXCLUDED.osm_mosm_kg, EXCLUDED.bioht_arabinose_mg_l, EXCLUDED.af_added_ml, EXCLUDED. af_totalizer_ml, EXCLUDED.theoretical_amount_of_feed_added_ml, EXCLUDED.percent_diff_in_feed, EXCLUDED.online_ph, EXCLUDED.broth_viscosity_ranking, EXCLUDED.bioht_pyruvate_mmol, EXCLUDED. total_amount_of_antifoam_added_ml, EXCLUDED.bioht_ldh_u_l, EXCLUDED.bioht_igg_mg_l, EXCLUDED.bioht_total_protein_g_l, EXCLUDED.sodium, EXCLUDED.potassium, EXCLUDED.bioht_formate_mg_l, EXCLUDED. percent_co2, EXCLUDED.bioht_glycerol_mg_l, EXCLUDED.datetime, EXCLUDED.gel_titer_g_l, EXCLUDED.antifoam_weight_g, EXCLUDED.scale_1_weight_g, EXCLUDED.pump_1_rpm, EXCLUDED.rate_from_scale_1_g_h, EXCLUDED.scale_2_weight_g, EXCLUDED.pump_2_rpm, EXCLUDED.rate_from_scale_2_g_h)", 
                        {"run_id":run_id,"ferm_sample_id":ferm_sample_id,"reactor_id_temp":reactor_id_temp,"ferm_stage":ferm_stage,"notes":notes,"eft_h":eft_h,"time_post_induction_h":time_post_induction_h,"rounded_post_induction_time_h":rounded_post_induction_time_h,"post_feed_start_time_h":post_feed_start_time_h,"o2_flow_percent":o2_flow_percent,"total_gas_flow_slpm":total_gas_flow_slpm,"base_totalizer_ml":base_totalizer_ml,"base_consumed_g":base_consumed_g,"feed_weight_g":feed_weight_g,"amount_of_feed_added_g":amount_of_feed_added_g,"amount_of_feed_added_ml":amount_of_feed_added_ml,"offline_ph":offline_ph,"od595":od595,"od_bioht":od_bioht,"bioht_glucose_g_l":bioht_glucose_g_l,"bioht_lactate_mmol":bioht_lactate_mmol,"bioht_acetate_mmol":bioht_acetate_mmol,"bioht_ammonia_mmol":bioht_ammonia_mmol,"bioht_glutamine_mmol":bioht_glutamine_mmol,"bioht_glutamate_mmol":bioht_glutamate_mmol,"bioht_phosphate_mmol":bioht_phosphate_mmol,"bioht_magnesium_mmol":bioht_magnesium_mmol,"osm_mosm_kg":osm_mosm_kg,"bioht_arabinose_mg_l":bioht_arabinose_mg_l,"af_added_ml":af_added_ml,"af_totalizer_ml":af_totalizer_ml,"theoretical_amount_of_feed_added_ml":theoretical_amount_of_feed_added_ml,"percent_diff_in_feed":percent_diff_in_feed,"online_ph":online_ph,"broth_viscosity_ranking":broth_viscosity_ranking,"bioht_pyruvate_mmol":bioht_pyruvate_mmol,"total_amount_of_antifoam_added_ml":total_amount_of_antifoam_added_ml,"bioht_ldh_u_l":bioht_ldh_u_l,"bioht_igg_mg_l":bioht_igg_mg_l,"bioht_total_protein_g_l":bioht_total_protein_g_l,"sodium":sodium,"potassium":potassium,"bioht_formate_mg_l":bioht_formate_mg_l,"percent_co2":percent_co2,"bioht_glycerol_mg_l":bioht_glycerol_mg_l,"datetime":datetime,"gel_titer_g_l":gel_titer_g_l,"last_id":last_ids[i],"antifoam_weight_g":antifoam_weight_g,"scale_1_weight_g":scale_1_weight_g,"pump_1_rpm":pump_1_rpm,"rate_from_scale_1_g_h":rate_from_scale_1_g_h,"scale_2_weight_g":scale_2_weight_g,"pump_2_rpm":pump_2_rpm,"rate_from_scale_2_g_h":rate_from_scale_2_g_h}) 
            conn.commit()
    return (render_template('submitted.html')) 
                #Response(
               # df.to_csv(),
                #mimetype="text/csv",
                #headers={"Content-disposition":
                #"attachment; filename=filename.csv"})

@app.route('/select_runs_summary', methods = ['GET', 'POST'])
def select_runs_summary():
    db.execute("SELECT * FROM pd_upstream.mfcs_online_data_file_list")
    mfcs_file_list_rows =  db.fetchall()
    run_data = []
    for row in mfcs_file_list_rows:
        run_data.append(row[0])
    
    
    db.execute("SELECT run_name FROM pd_upstream.sutro_production_batch_record order by id desc")
    run_name_rows = db.fetchall()
    run_names_lst = []
    for r in run_name_rows:
        run_names_lst.append(r[0])

    insp = inspect(engine)
    col_tables = insp.get_columns('growth_data','pd_upstream')
    growth_fields = []
    for c in col_tables:
        growth_fields.append(c['name'])
    
    
    conn.commit()
    return render_template('select_runs_summary.html',run_names_lst=run_names_lst,run_data=run_data, growth_fields = growth_fields)    

@app.route('/select_runs_summary_download', methods = ['POST'])
def select_runs_summary_download():
    run_ids=request.form.getlist("run_ids")
    run_ids_tup = tuple(run_ids)

    insp = inspect(engine)
    col_tables = insp.get_columns('growth_data','pd_upstream')
    all_growth_fields = []
    for c in col_tables:
        all_growth_fields.append(c['name'])

    selected_growth_flds = request.form.getlist("growth_fields")
    #try:
    #    db.execute("SELECT run_name,scale_liter from pd_upstream.sutro_production_batch_record WHERE run_name IN %s",(run_ids_tup,))
    #except: 
    #    conn.rollback()
    #df = pd.DataFrame(db.fetchall(), columns = ['run_name','scale_liter'])
    query = """SELECT * from pd_upstream.growth_data WHERE run_id IN :run_list"""
    params = {'run_list':run_ids,}
    t = text(query)
    t = t.bindparams(bindparam('run_list', expanding = True))
    #connection.execute(t,params)
    df = pd.DataFrame(connection.execute(t,params), columns = all_growth_fields)
    df=df[selected_growth_flds]
    session["df"] = df.to_csv(index=False)


     # Get the CSV data as a string from the session
    csv = session["df"] if "df" in session else ""
    
    # Create a string buffer
    buf_str = io.StringIO(csv)

    # Create a bytes buffer from the string buffer
    buf_byt = io.BytesIO(buf_str.read().encode("utf-8"))
    
    #conn.commit()
    #trans.commit()
    # Return the CSV data as an attachment

    return send_file(buf_byt,
                     mimetype="text/csv",
                     as_attachment=True,
                     download_name="data.csv")

@app.route('/select_runs_online_br_spotfire', methods = ['GET', 'POST'])
def select_runs_online_br_spotfire():
    try:
        trans = connection.begin()

        mfcs_file_list_rows = connection.execute(text("SELECT * FROM pd_upstream.mfcs_online_data_file_list"))
        #mfcs_file_list_rows =  connection.fetchall()
        run_data = []
        for row in mfcs_file_list_rows:
            run_data.append(row[0])
        
        
        run_name_rows = connection.execute(text("SELECT run_name FROM pd_upstream.sutro_production_batch_record order by id desc"))
        #run_name_rows = connection.fetchall()
        run_names_lst = []
        for r in run_name_rows:
            run_names_lst.append(r[0])
        trans.commit()
    except Exception:
        trans.rollback()
        raise

    return render_template('select_runs_online_br_spotfire.html',run_names_lst=run_names_lst,run_data=run_data)

@app.route('/select_runs_online_br_spotfire_submit', methods=['GET','POST'])
def select_runs_online_br_spotfire_submit():
    run_ids= request.form.getlist("run_id_1")
    run_ids_tup = tuple(run_ids)
    try:
        trans = connection.begin()

        query = """create or replace view pd_upstream.mfcs_spotfire_selectruns as \
            select view2.*, EXTRACT(EPOCH FROM (date_time - actual_inoculation_time))/3600 AS eft_h_mfcs, EXTRACT(EPOCH FROM (date_time - actual_induction_start_time))/3600 AS time_post_induction_h_mfcs \
            from \
                (select view1.*, spbr.*, sh.percent_solid_pre_wash, a.titer, a.activity\
                from\
                    (select coalesce(growth_data.datetime, pd_upstream.mfcs_qs.pdattime) as date_time, coalesce(growth_data.run_id, mfcs_qs.run_id_m) as run_id_1, growth_data.*, mfcs_qs.*\
                    from pd_upstream.mfcs_qs\
                    full outer join pd_upstream.growth_data on pd_upstream.mfcs_qs.pdattime=pd_upstream.growth_data.datetime and pd_upstream.mfcs_qs.run_id_m=pd_upstream.growth_data.run_id) view1\
                left join pd_upstream.sutro_production_batch_record spbr on spbr.run_name = view1.run_id_1\
                left join pd_upstream.smallscale_harvest sh on sh.ferm_sample_id = view1.ferm_sample_id\
                left join pd_upstream.activity_titer a on a.ferm_sample_id = view1.ferm_sample_id) view2\
            WHERE run_id_1 IN :run_id"""
        params = {'run_id':run_ids,}

        t = text(query)
        t = t.bindparams(bindparam('run_id', expanding = True))
        connection.execute(t,params)
        
        trans.commit()
    except Exception:
        trans.rollback()
        raise


    trans = connection.begin()

    insp = inspect(engine)
    col_tables = insp.get_columns('mfcs_spotfire_selectruns','pd_upstream')
    all_view_fields = []
    for c in col_tables:
        all_view_fields.append(c['name'])
    query_select = """select * FROM pd_upstream.mfcs_spotfire_selectruns"""
    t = text(query_select)
    df = pd.DataFrame(connection.execute(t), columns = all_view_fields)
    df = df.head()
    session["df"] = df.to_csv(index=False)
    trans.commit()

    return (render_template('select_runs_online_br_spotfire_submit.html') )

#https://stackoverflow.com/questions/62823361/download-dataframe-as-csv-from-flask
@app.route("/select_runs_online_br_spotfire_download", methods=['POST'])
def select_runs_online_br_spotfire_download():

  # Get the CSV data as a string from the session
    csv = session["df"] if "df" in session else ""
    
    # Create a string buffer
    buf_str = io.StringIO(csv)

    # Create a bytes buffer from the string buffer
    buf_byt = io.BytesIO(buf_str.read().encode("utf-8"))
    
    # Return the CSV data as an attachment
    return send_file(buf_byt,
                     mimetype="text/csv",
                     as_attachment=True,
                     download_name="data.csv")
   
@app.route('/upload_online_bioreactor_data', methods=['GET','POST'])
def upload_online_bioreactor_data():
   
    db.execute("SELECT * FROM pd_upstream.mfcs_online_data_file_list")
    mfcs_file_list_rows =  db.fetchall()
    run_data = []
    for row in mfcs_file_list_rows:
        run_data.append(row[0])

    db.execute("SELECT run_name FROM pd_upstream.sutro_production_batch_record order by id desc")
    run_name_rows = db.fetchall()
    run_names_lst = []
    for r in run_name_rows:
        run_names_lst.append(r[0])
 
    return render_template('upload_online_bioreactor_data.html',run_data=run_data,run_names_lst=run_names_lst)

@app.route('/upload_online_bioreactor_data_submitted', methods=['GET','POST'])
def upload_online_bioreactor_data_submitted():
    def compare_df_db_cols(df,db_flds_set): #function to compare df fields and db fields
            dfcollist = list(df.columns) #store columns of dataframe into a list
            dfcolset = set(dfcollist)
             # comparee db fields cols and excel dataframe cols and find what is not in the the db fields list
            #dfcolset.difference(db_flds_set)
            setdifflst = list(dfcolset.difference(db_flds_set))
            db_flds_set -= dfcolset.difference(db_flds_set) # remove the differences in the db fields set and dfcolset from the db fields set
           # dbfieldssetlist = list(db_flds_set)
            return setdifflst

    def convert_df_to_lst(excel_df,db_fld_lst): #function to convert excel_df to a list where it is aligned to list of database fields db_fld_lst
            excel_df = excel_df.dropna(how='all') #drop all rows that are completely blank in Excel file to be imported
            db_flds_set = set(db_fld_lst) #convert db_flds_lst to a set
            setdifflst=compare_df_db_cols(df=excel_df, db_flds_set=db_flds_set) #function returns variable 'setdifflst' to see what fields in df from excel file are not in db
            excel_df.drop(setdifflst, axis=1, inplace=True)
            excel_df = excel_df.reindex(columns=db_fld_lst)     #realign dataframe to database fields order
            excel_df =excel_df.replace({np.nan: None}) 
            return excel_df

    bioreactor_system = request.form.get('bioreactor-system')
    bioreactor_system = str(bioreactor_system)

    if bioreactor_system == 'D30-MFCS2':
        db_name_to_excel_name = {
                "PDatTime":"pdattime",
                "Age (h)":"age_h",
                "TEMP_1_Value":"temp_value",
                "_BatchAlarmCount_Value":"_batchalarmcount_value",
                "_BatchSync_Value":"_batchsync_value",
                "pHA_1_Value":"ph_value",
                "pO2A_1_Value":"po2_value",
                "STIRR_1_Value":"stirr_value",
                "_BatchAge_Value":"_batchage_value",
                "ACIDT_Value":"acidt_value",
                "BASET_Value":"baset_value",
                "JTEMP_Value":"jtemp_value",
                "O2EN_Value":"o2en_value",
                "O2EN_Setpoint":"o2en_setpoint",
                "pH_Setpoint":'ph_setpoint',
                "pO2_Setpoint":"po2_setpoint",
                "STIRR_Setpoint":"stirr_setpoint",
                "SUBSA_Value":"subsa_value",
                "SUBSA_Setpoint":"subsa_setpoint",
                "SUBSB_Value":"subsb_value",
                "SUBSB_Setpoint":"subsb_setpoint",
                "TEMP_Setpoint":"temp_setpoint", 
                "AIRSP_1_Value":"airflow_value_slpm",
                "O2SP_1_Value":"o2flow_value_slpm",
                "PRESS_1_Value":"pressure_value_psi"
                }
    elif bioreactor_system == 'Qs-MFCS2':
        db_name_to_excel_name = {
                "PDatTime":"pdattime",
                "Age (h)":"age_h",
                "TEMP_Value":"temp_value",
                "_BatchAlarmCount_Value":"_batchalarmcount_value",
                "_BatchSync_Value":"_batchsync_value",
                "pH_Value":"ph_value",
                "pO2_Value":"po2_value",
                "STIRR_Value":"stirr_value",
                "_BatchAge_Value":"_batchage_value",
                "ACIDT_Value":"acidt_value",
                "BASET_Value":"baset_value",
                "EXT_Value":"ext_value",
                "FO_LE_T_Value":"fo_le_t_value",
                "FO_LE_T_Setpoint":"fo_le_t_setpoint",
                "GASFL_Value":"gasfl_value",
                "GASFL_Setpoint":"gasfl_setpoint",
                "JTEMP_Value":"jtemp_value",
                "O2EN_Value":"o2en_value",
                "O2EN_Setpoint":"o2en_setpoint",
                "pH_Setpoint":'ph_setpoint',
                "pO2_Setpoint":"po2_setpoint",
                "STIRR_Setpoint":"stirr_setpoint",
                "SUBSA_Value":"subsa_value",
                "SUBSA_Setpoint":"subsa_setpoint",
                "SUBSB_Value":"subsb_value",
                "SUBSB_Setpoint":"subsb_setpoint",
                "TEMP_Setpoint":"temp_setpoint" 
                }
    elif bioreactor_system =='1L-Bs-MFCS4':
        db_name_to_excel_name = {
                "PDatTime Unnamed: 1_level_1":"pdattime",
                "ProcessTime Value":"age_h",
                "TEMP Value":"temp_value",
                "pH Value":"ph_value",
                "pO2 Value":"po2_value",
                "STIRR Value":"stirr_value",
                "ACIDT Value":"acidt_value",
                "BASET Value":"baset_value",
                "JTEMP Value":"jtemp_value",
                "pH Setpoint":'ph_setpoint',
                "pO2 Setpoint":"po2_setpoint",
                "STIRR Setpoint":"stirr_setpoint",
                "SUBS_A Value":"subsa_value",
                "SUBS_A Setpoint":"subsa_setpoint",
                "SUBS_B Value":"subsb_value",
                "SUBS_B Setpoint":"subsb_setpoint",
                "TEMP Setpoint":"temp_setpoint",
                "AIRSP_A Value":"airflow_value_slpm",
                "O2SP_A Value":"o2flow_value_slpm"
                }
   
    num_runs = request.form.get('num_runs')
    # Each file will take about ~ 1.5 minutes to upload (for 65k rows in the Excel file), 3 minutes if run id already in mfcs_qs
    for run in range(int(num_runs)):
        run_id = request.form.get('run_id_' + str(run))
        #st = "SELECT id FROM pd_upstream.sutro_production_batch_record WHERE run_name='{}'".format(run_id)
        db.execute("SELECT id from pd_upstream.sutro_production_batch_record WHERE run_name = %(run_id)s", {"run_id":run_id})
        #actual_run_id_query = db1.execute(st, run_id)
        actual_run_id_query = db.fetchall()
        actual_run_id = 0
        #for i in actual_run_id_query:
        #    actual_run_id = int(i[0])
        for i in actual_run_id_query:
           actual_run_id=int(i[0])
        # associates with the most recently added row (to sutro production batch record) of this run id (but ideally the same run id shouldn't be in there more than once)
        
        #s = "DELETE FROM pd_upstream.mfcs_qs WHERE run_id='{}';".format(run_id)
        db.execute("DELETE FROM pd_upstream.mfcs_qs WHERE run_id_m= %(run_id_m)s", {"run_id_m":run_id})
        #db1.execute(s)
        #db1.commit()
        conn.commit()
        f = request.files['file_' + str(run)]
        filename = f.filename
        #q = "DELETE FROM pd_upstream.mfcs_online_data_file_list WHERE run_id='{}';".format(run_id)
        #db1.execute(q)
        db.execute("DELETE FROM pd_upstream.mfcs_online_data_file_list WHERE run_id= %(run_id)s", {"run_id":run_id})
        #db1.commit()
        conn.commit()
        #db1.execute("INSERT INTO pd_upstream.mfcs_online_data_file_list (filename,run_id) VALUES (:filename,:run_id)",{"filename":filename,"run_id":run_id})
        #db1.commit()
        db.execute("""INSERT INTO pd_upstream.mfcs_online_data_file_list (filename,run_id) VALUES (%(filename)s,%(run_id)s)""",{"filename":filename,"run_id":run_id})

        conn.commit()
        if bioreactor_system =='D30-MFCS2' or bioreactor_system == 'Qs-MFCS2':
            df = pd.read_excel(f)
        elif bioreactor_system == '1L-Bs-MFCS4':
            df = pd.read_csv(f,sep=";",header=[0,1])
            df.columns = [f'{i} {j}' for i,j in df.columns]
            df = df.drop(labels=0, axis=0)
      
            df=df.replace('UNCERTAIN', np.NaN)
        df['filename'] = filename
        df['actual_run_id_m'] = actual_run_id
        df['run_id_m'] = run_id

        #rename fields in df 
        df= df.rename(columns = db_name_to_excel_name)
        df['pdattime'] = pd.to_datetime(df['pdattime'])

        insp = inspect(engine)
        columns_table = insp.get_columns('mfcs_qs', 'pd_upstream') #schema is optional
        fields = []
        for c in columns_table :
            fields.append(c['name'])

        df=convert_df_to_lst(df,fields)
       
        df.to_sql(name ='mfcs_qs', index = False, con=engine, if_exists='append', method = 'multi', schema = 'pd_upstream', chunksize = 1000)
        
        engine.dispose()

    #return render_template('submitted.html', diff_lst = online_br_diff_lst )
    return render_template('submitted.html')


@app.route('/upload_ferm_batch_record')    
def upload_ferm_batch_record():

    return render_template('upload_ferm_batch_record.html')


@app.route('/upload_ferm_batch_record_final', methods = ["POST"])    
def upload_ferm_batch_record_final():

    def compare_df_db_cols(df,db_flds_set): #function to compare df fields and db fields
            dfcollist = list(df.columns) #store columns of dataframe into a list
            dfcolset = set(dfcollist)
            setdifflst = list(dfcolset.difference(db_flds_set)) #comparee db fields cols and excel dataframe cols and find what is not in the the db fields list
            db_flds_set -= dfcolset.difference(db_flds_set) # remove the differences in the db fields set and dfcolset from the db fields set
            return setdifflst
    
    def convert_df_to_lst(excel_df,db_fld_lst): #function to convert excel_df to a list where it is aligned to list of database fields db_fld_lst
            excel_df = excel_df.dropna(how='all') #drop all rows that are completely blank in Excel file to be imported
            db_flds_set = set(db_fld_lst) #convert db_flds_lst to a set
            setdifflst=compare_df_db_cols(df=excel_df, db_flds_set=db_flds_set) #function returns variable 'setdifflst' to see what fields in df from excel file are not in db
            excel_df.drop(setdifflst, axis=1, inplace=True)
            excel_df = excel_df.reindex(columns=db_fld_lst) #realign dataframe to database fields order
            excel_df =excel_df.replace({np.nan: None}) 
            dflist = excel_df.values.tolist()
            return setdifflst,dflist


    #insert sutro productio batch record tab
    try:
        experiment_name = request.form.get('experiment_name')
        experiment_name = str(experiment_name)
        today = date.today()
        date_1 = today.strftime("%Y-%m-%d")
        exp_lst = [experiment_name, date_1]
        insert_sql = ''' INSERT INTO pd_upstream.experiment(experiment_name, date1) VALUES (%s,%s) ON CONFLICT (experiment_name) DO UPDATE SET (experiment_name,date1) =(EXCLUDED.experiment_name, EXCLUDED.date1)'''
        db.execute(insert_sql, exp_lst)
        conn.commit()
        ferm_file = request.files['file_br']
        #insert sutro productio batch record tab
        def rmv_vals_frm_lst_br(lst,val1):
            lst.remove(val1)

        db.execute('select * from sutro_production_batch_record') #select all from GrowthData table
        dbfieldslist = [description[0] for description in db.description] # store fields of database into a list
        del dbfieldslist[0] # delete first index of list that contains fields of database, specifically id
        db_flds_set_br = set(dbfieldslist) 
        wb = load_workbook(ferm_file)
        sh = wb["General Process Conditions"]
        rowcounterlst = []
        for row in sh.iter_rows():
            print(row[0].value)
            color_in_hex = row[0].fill.start_color.index
            if color_in_hex == 4:
                rowcounterlst.append(str(row[0].row))
        rowcounterlst = list(map(int, rowcounterlst))
        rowcounterlst[:] = [number - 1 for number in rowcounterlst] #subtract 1 from each list
        df_br = pd.read_excel(ferm_file, 'General Process Conditions', header = None)
        df_br=df_br.drop(rowcounterlst) # drop original header row
        df_br = df_br.T # transpose dataframe
        df_br=df_br.rename(columns=df_br.iloc[0]).drop(df_br.index[0])
        df_br = df_br.reset_index(drop=True) #reset index
        df_br = df_br.dropna(how='all') #drop all rows that are completely blank
        select_id_exp = '''SELECT id from pd_upstream.experiment
                            WHERE experiment_name = ANY(%(parameter_array)s);'''
        db.execute(select_id_exp,{"parameter_array":[experiment_name]})
        exp_ids = db.fetchall() # returns tuple
        exp_ids_lst = list(exp_ids[0])
        df_br["experiment_id"]=exp_ids_lst[0]
        br_setdifflst=compare_df_db_cols(df=df_br, db_flds_set = db_flds_set_br) #function returns variable 'setdifflst' to see what fields in df are not in db
        df_br.drop(br_setdifflst, axis=1, inplace=True)
        df_br = df_br.reindex(columns=dbfieldslist) #align columns to MFCS table in DB
        df_br = df_br.replace({np.nan: None})
        df_br.replace({pd.NaT: None}, inplace=True)
        dflist_br = df_br.values.tolist()
        #take out 1 less the number %s placeholders generated by the placeholder script
        insert_sql = '''
            INSERT INTO sutro_production_batch_record (process_version_name, run_name, reactor_id, run_description, product, strain, site_of_run, scale_liter, number_of_seed_stages, seed_train_run_id, feed_medium_concentration, batch_temperature_setpoint_celcius, fed_batch_temperature_setpoint_celcius, induction_temperature_setpoint_celcius, fed_batch_ph_setpoint, fed_batch_do_setpoint, target_batch_phase_airflow_vvm, target_fed_batch_phase_airflow_vvm, foam_control, inducer, target_batch_volume_ml, target_pre_induction_volume_ml, target_final_volume_ml, target_final_volume_w_drawdown_ml, seed_media, feed_media_1, target_time_to_add_feed_media_1_post_feed_start_h, feed_media_2, target_time_to_add_feed_media_2_post_feed_start_h, estimated_batch_timing_h, target_post_feed_duration_until_temp_shift_h, target_post_feed_duration_until_induction_h, target_induction_duration_h, growth_feed_rate_1_m_h_1, growth_feed_rate_1_post_feed_target_time_h, growth_feed_rate_1_type, induction_feed_rate_m_h_1, post_feed_time_of_percent_drawdown_h, drawdown_at_post_feed_time_percent, initial_batch_vol_litter, x0_grams_over_litter, sf_feed_gluc_grams_over_litter, s0_batch_gluc_grams_over_litter, yx_over_s_grams_over_grams, od_over_gram_dry_wt_over_litter, growth_phase_feed_interval_h, feed_profile_type, induction_phase_feed_interval_h, induction_feed_profile_type, actual_inoculation_time, actual_fed_batch_start_time, actual_induction_start_time, target_induction_duration_1_h, growth_feed_rate_2_m_h_1, growth_feed_rate_2_post_feed_target_time_h, growth_feed_rate_2_type, run_outcome, batch_ph_setpoint, induction_ph_setpoint, experiment_description, feed_profile_version, experiment_id,process_version_name_used_as_template, feed_media_1_density, addition_on_scale_pump_1, addition_on_scale_pump_2, batch_phase_ph_setpoint, batch_phase_do_setpoint, induction_do_setpoint, process_description, max_tip_speed_ms, max_stirr_rpm,run_alias,experiment_name,campaign_name,experiment_purpose,impeller_diameter_m)
            VALUES (%s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (run_name) 
            DO UPDATE SET (process_version_name, run_name, reactor_id, run_description, product, strain, site_of_run, scale_liter, number_of_seed_stages, seed_train_run_id, feed_medium_concentration, batch_temperature_setpoint_celcius, fed_batch_temperature_setpoint_celcius, induction_temperature_setpoint_celcius, fed_batch_ph_setpoint, fed_batch_do_setpoint, target_batch_phase_airflow_vvm, target_fed_batch_phase_airflow_vvm, foam_control, inducer, target_batch_volume_ml, target_pre_induction_volume_ml, target_final_volume_ml, target_final_volume_w_drawdown_ml, seed_media, feed_media_1, target_time_to_add_feed_media_1_post_feed_start_h, feed_media_2, target_time_to_add_feed_media_2_post_feed_start_h, estimated_batch_timing_h, target_post_feed_duration_until_temp_shift_h, target_post_feed_duration_until_induction_h, target_induction_duration_h, growth_feed_rate_1_m_h_1, growth_feed_rate_1_post_feed_target_time_h, growth_feed_rate_1_type, induction_feed_rate_m_h_1, post_feed_time_of_percent_drawdown_h, drawdown_at_post_feed_time_percent, initial_batch_vol_litter, x0_grams_over_litter, sf_feed_gluc_grams_over_litter, s0_batch_gluc_grams_over_litter, yx_over_s_grams_over_grams, od_over_gram_dry_wt_over_litter, growth_phase_feed_interval_h, feed_profile_type, induction_phase_feed_interval_h, induction_feed_profile_type, actual_inoculation_time, actual_fed_batch_start_time, actual_induction_start_time, target_induction_duration_1_h, growth_feed_rate_2_m_h_1, growth_feed_rate_2_post_feed_target_time_h, growth_feed_rate_2_type, run_outcome, batch_ph_setpoint, induction_ph_setpoint, experiment_description, feed_profile_version, experiment_id, process_version_name_used_as_template, feed_media_1_density, addition_on_scale_pump_1, addition_on_scale_pump_2, batch_phase_ph_setpoint, batch_phase_do_setpoint, induction_do_setpoint, process_description, max_tip_speed_ms, max_stirr_rpm,run_alias,experiment_name,campaign_name,experiment_purpose,impeller_diameter_m) = (EXCLUDED.process_version_name, EXCLUDED.run_name, EXCLUDED.reactor_id, EXCLUDED.run_description, EXCLUDED.product, EXCLUDED.strain, EXCLUDED.site_of_run, EXCLUDED.scale_liter, EXCLUDED.number_of_seed_stages, EXCLUDED.seed_train_run_id, EXCLUDED.feed_medium_concentration, EXCLUDED.batch_temperature_setpoint_celcius, EXCLUDED.fed_batch_temperature_setpoint_celcius, EXCLUDED.induction_temperature_setpoint_celcius, EXCLUDED.fed_batch_ph_setpoint, EXCLUDED.fed_batch_do_setpoint, EXCLUDED.target_batch_phase_airflow_vvm, EXCLUDED.target_fed_batch_phase_airflow_vvm, EXCLUDED.foam_control, EXCLUDED.inducer, EXCLUDED.target_batch_volume_ml, EXCLUDED.target_pre_induction_volume_ml, EXCLUDED.target_final_volume_ml, EXCLUDED.target_final_volume_w_drawdown_ml, EXCLUDED.seed_media, EXCLUDED.feed_media_1, EXCLUDED.target_time_to_add_feed_media_1_post_feed_start_h, EXCLUDED.feed_media_2, EXCLUDED.target_time_to_add_feed_media_2_post_feed_start_h, EXCLUDED.estimated_batch_timing_h, EXCLUDED.target_post_feed_duration_until_temp_shift_h, EXCLUDED.target_post_feed_duration_until_induction_h, EXCLUDED.target_induction_duration_h, EXCLUDED.growth_feed_rate_1_m_h_1, EXCLUDED.growth_feed_rate_1_post_feed_target_time_h, EXCLUDED.growth_feed_rate_1_type, EXCLUDED.induction_feed_rate_m_h_1, EXCLUDED.post_feed_time_of_percent_drawdown_h, EXCLUDED.drawdown_at_post_feed_time_percent, EXCLUDED.initial_batch_vol_litter, EXCLUDED.x0_grams_over_litter, EXCLUDED.sf_feed_gluc_grams_over_litter, EXCLUDED.s0_batch_gluc_grams_over_litter, EXCLUDED.yx_over_s_grams_over_grams, EXCLUDED.od_over_gram_dry_wt_over_litter, EXCLUDED.growth_phase_feed_interval_h, EXCLUDED.feed_profile_type, EXCLUDED.induction_phase_feed_interval_h, EXCLUDED.induction_feed_profile_type, EXCLUDED.actual_inoculation_time, EXCLUDED.actual_fed_batch_start_time, EXCLUDED.actual_induction_start_time, EXCLUDED.target_induction_duration_1_h, EXCLUDED.growth_feed_rate_2_m_h_1, EXCLUDED.growth_feed_rate_2_post_feed_target_time_h, EXCLUDED.growth_feed_rate_2_type, EXCLUDED.run_outcome, EXCLUDED.batch_ph_setpoint, EXCLUDED.induction_ph_setpoint, EXCLUDED.experiment_description, EXCLUDED.feed_profile_version, EXCLUDED.experiment_id, EXCLUDED.process_version_name_used_as_template, EXCLUDED.feed_media_1_density, EXCLUDED.addition_on_scale_pump_1, EXCLUDED.addition_on_scale_pump_2, EXCLUDED.batch_phase_ph_setpoint, EXCLUDED.batch_phase_do_setpoint, EXCLUDED.induction_do_setpoint, EXCLUDED.process_description, EXCLUDED.max_tip_speed_ms, EXCLUDED.max_stirr_rpm,EXCLUDED.run_alias,EXCLUDED.experiment_name,EXCLUDED.campaign_name,EXCLUDED.experiment_purpose,EXCLUDED.impeller_diameter_m) 
            '''
        db.executemany(insert_sql, dflist_br)
        
         #insert growth_data tab
        def rmv_vals_frm_lst(lst,val1):
            lst.remove(val1)
        db.execute('select * from growth_data') #select all from GrowthData database table
        dbfieldslist = [description[0] for description in db.description] # store fields of database into a list
        val1 = 'id_g'
        rmv_vals_frm_lst(dbfieldslist,val1)
        df_growth = pd.read_excel(ferm_file, 'Growth Data') #read Excel file to be imported
        df_growth = df_growth.dropna(how='all') #drop all rows that are completely blank in Excel file to be imported
        runid_list = df_growth["run_id"].values.tolist()
        #get id of run_name from sutro_production_batch_record
        select_id_sql = '''
                    SELECT run_name,id from sutro_production_batch_record 
                    WHERE run_name = ANY(%s);
                    '''
        db.execute(select_id_sql, (runid_list,))
        allids = db.fetchall()
        listallids = list(allids)
        dfidlist = pd.DataFrame(listallids, columns =['run_id','actual_run_id'])
        df_growth=df_growth.merge(dfidlist, on='run_id', how='left')
        df_growth['datetime'] = df_growth['datetime'].astype(str)
        df_growth['datetime'] = pd.to_datetime(df_growth['datetime'])
        df_growth['datetime'] = df_growth['datetime'].apply(lambda x: x.strftime('%Y-%m-%d, %H:%M:%S')if not pd.isnull(x) else None)
        growth_setdifflst, dflist = convert_df_to_lst(df_growth,dbfieldslist)
        #dflist = df_growth.values.tolist()
    
        for i in range(len(dflist)):
            print(dflist[i])
        insert_sql = '''
            INSERT INTO growth_data (run_id, ferm_sample_id, reactor_id_temp, ferm_stage, notes, eft_h, time_post_induction_h, post_feed_start_time_h, o2_flow_percent, total_gas_flow_slpm, base_totalizer_ml, base_consumed_g, feed_weight_g, amount_of_feed_added_g, amount_of_feed_added_ml, offline_ph, od595, od_bioht, bioht_glucose_g_l, bioht_lactate_mmol, bioht_acetate_mmol, bioht_ammonia_mmol, bioht_glutamine_mmol, bioht_glutamate_mmol, bioht_phosphate_mmol, bioht_magnesium_mmol, osm_mosm_kg, bioht_arabinose_mg_l, af_added_ml, af_totalizer_ml, theoretical_amount_of_feed_added_ml, percent_diff_in_feed, online_ph, broth_viscosity_ranking, bioht_pyruvate_mmol, total_amount_of_antifoam_added_ml, bioht_ldh_u_l, bioht_igg_mg_l, bioht_total_protein_g_l, sodium, potassium, bioht_formate_mg_l, percent_co2, bioht_glycerol_mg_l, datetime, gel_titer_g_l, actual_run_id, pump_1_rpm, pump_2_rpm, rate_from_scale_1_g_h, rate_from_scale_2_g_h, scale_1_weight_g, scale_2_weight_g, antifoam_weight_g)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (run_id,ferm_sample_id) DO UPDATE SET
            (run_id, ferm_sample_id, reactor_id_temp, ferm_stage, notes, eft_h, time_post_induction_h, post_feed_start_time_h, o2_flow_percent, total_gas_flow_slpm, base_totalizer_ml, base_consumed_g, feed_weight_g, amount_of_feed_added_g, amount_of_feed_added_ml, offline_ph, od595, od_bioht, bioht_glucose_g_l, bioht_lactate_mmol, bioht_acetate_mmol, bioht_ammonia_mmol, bioht_glutamine_mmol, bioht_glutamate_mmol, bioht_phosphate_mmol, bioht_magnesium_mmol, osm_mosm_kg, bioht_arabinose_mg_l, af_added_ml, af_totalizer_ml, theoretical_amount_of_feed_added_ml, percent_diff_in_feed, online_ph, broth_viscosity_ranking, bioht_pyruvate_mmol, total_amount_of_antifoam_added_ml, bioht_ldh_u_l, bioht_igg_mg_l, bioht_total_protein_g_l, sodium, potassium, bioht_formate_mg_l, percent_co2, bioht_glycerol_mg_l, datetime, gel_titer_g_l, actual_run_id, pump_1_rpm, pump_2_rpm, rate_from_scale_1_g_h, rate_from_scale_2_g_h, scale_1_weight_g, scale_2_weight_g, antifoam_weight_g) =
            (EXCLUDED.run_id, EXCLUDED.ferm_sample_id, EXCLUDED.reactor_id_temp, EXCLUDED.ferm_stage, EXCLUDED.notes, EXCLUDED.eft_h, EXCLUDED.time_post_induction_h, EXCLUDED.post_feed_start_time_h, EXCLUDED.o2_flow_percent, EXCLUDED.total_gas_flow_slpm, EXCLUDED.base_totalizer_ml, EXCLUDED.base_consumed_g, EXCLUDED.feed_weight_g, EXCLUDED.amount_of_feed_added_g, EXCLUDED.amount_of_feed_added_ml, EXCLUDED.offline_ph, EXCLUDED.od595, EXCLUDED.od_bioht, EXCLUDED.bioht_glucose_g_l, EXCLUDED.bioht_lactate_mmol, EXCLUDED.bioht_acetate_mmol, EXCLUDED.bioht_ammonia_mmol, EXCLUDED.bioht_glutamine_mmol, EXCLUDED.bioht_glutamate_mmol, EXCLUDED.bioht_phosphate_mmol, EXCLUDED.bioht_magnesium_mmol, EXCLUDED.osm_mosm_kg, EXCLUDED.bioht_arabinose_mg_l, EXCLUDED.af_added_ml, EXCLUDED.af_totalizer_ml, EXCLUDED.theoretical_amount_of_feed_added_ml, EXCLUDED.percent_diff_in_feed, EXCLUDED.online_ph, EXCLUDED.broth_viscosity_ranking, EXCLUDED.bioht_pyruvate_mmol, EXCLUDED.total_amount_of_antifoam_added_ml, EXCLUDED.bioht_ldh_u_l, EXCLUDED.bioht_igg_mg_l, EXCLUDED.bioht_total_protein_g_l, EXCLUDED.sodium, EXCLUDED.potassium, EXCLUDED.bioht_formate_mg_l, EXCLUDED.percent_co2, EXCLUDED.bioht_glycerol_mg_l, EXCLUDED.datetime, EXCLUDED.gel_titer_g_l, EXCLUDED.actual_run_id,EXCLUDED.pump_1_rpm, EXCLUDED.pump_2_rpm, EXCLUDED.rate_from_scale_1_g_h, EXCLUDED.rate_from_scale_2_g_h, EXCLUDED.scale_1_weight_g, EXCLUDED.scale_2_weight_g, EXCLUDED.antifoam_weight_g)     
        '''
        db.executemany(insert_sql, dflist)
        
        #insert activity_titer tab
        
        db.execute('select * from activity_titer') #select all from GrowthData database table
        db_fld_lst = [description[0] for description in db.description] # store fields of database into a list
        #val1 = 'id'
        #rmv_vals_frm_lst(activity_titer_list,val1)
        excel_df = pd.read_excel(ferm_file, 'Activity_Titer') #read Excel file to be imported
        activity_setdifflst, dflist = convert_df_to_lst(excel_df, db_fld_lst)
      

        insert_sql = '''
                INSERT INTO activity_titer (run_id_at, activity, activity_units, lysate_lot_id, ferm_sample_id, official_result, titer, titer_assay_type, titer_units, post_harvest_stage, cell_disruption_equipment, cell_disruption_eq_id, clarified_lysate_lot_id, bioit_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (run_id_at,ferm_sample_id) DO UPDATE SET
            (run_id_at,activity,activity_units,lysate_lot_id,ferm_sample_id,official_result,titer,titer_assay_type,titer_units,post_harvest_stage,cell_disruption_equipment,cell_disruption_eq_id,clarified_lysate_lot_id,bioit_id)=
            (EXCLUDED.run_id_at, EXCLUDED.activity, EXCLUDED.activity_units, EXCLUDED.lysate_lot_id, EXCLUDED.ferm_sample_id, EXCLUDED.official_result, EXCLUDED.titer, EXCLUDED.titer_assay_type, EXCLUDED.titer_units, EXCLUDED.post_harvest_stage, EXCLUDED.cell_disruption_equipment, EXCLUDED.cell_disruption_eq_id, EXCLUDED.clarified_lysate_lot_id, EXCLUDED.bioit_id)
            '''
        db.executemany(insert_sql, dflist)
    
    finally:
        conn.commit()
    return render_template('upload_ferm_batch_record_final.html',br_fld_lst=br_setdifflst, growth_fld_lst=growth_setdifflst, activity_fld_lst=activity_setdifflst)


@app.route('/batch_process_edit')
def batch_process_edit():
    try:
        trans = connection.begin()

        #db.execute("SELECT * FROM pd_upstream.default_batch_process")
        #dbp_rows = db.fetchall()
        dbp_rows = connection.execute(text("SELECT * FROM pd_upstream.default_batch_process"))
        insp = inspect(engine)
        tbl_cols = insp.get_columns('default_batch_process','pd_upstream')
        col_names = []
        for c in tbl_cols:
            col_names.append(c['name'])
        #rowkeys = [desc[0] for desc in db.description]
        rows = []

        for r in dbp_rows:
           rows.append(dict(zip(col_names, r)))
           
        processes = connection.execute(text("SELECT process_version_name_used_as_template FROM pd_upstream.default_batch_process order by process_version_name_used_as_template"))
        versions = []
        for v in processes:
            versions.append(v[0])
        trans.commit()
    except Exception:
        trans.rollback()
        raise
    return render_template('batch_process_edit.html', rows=rows, versions = versions)

@app.route('/batch_process_edit_submitted', methods=["POST"])
def batch_process_edit_submitted():
    process_version_name_used_as_template = request.form.get("process_version_name_used_as_template")
    process_description = request.form.get("process_description")
    strain = request.form.get("strain")
    scale_liter = request.form.get("scale_liter")
    number_of_seed_stages = request.form.get("number_of_seed_stages")
    feed_medium_concentration = request.form.get("feed_medium_concentration")
    batch_temperature_setpoint_celcius = request.form.get("batch_temperature_setpoint_celcius")
    fed_batch_temperature_setpoint_celcius = request.form.get("fed_batch_temperature_setpoint_celcius")
    induction_temperature_setpoint_celcius = request.form.get("induction_temperature_setpoint_celcius")
    batch_phase_ph_setpoint = request.form.get("batch_phase_ph_setpoint")
    fed_batch_ph_setpoint = request.form.get("fed_batch_ph_setpoint")
    induction_ph_setpoint = request.form.get("induction_ph_setpoint")
    batch_phase_do_setpoint = request.form.get("batch_phase_do_setpoint")
    fed_batch_do_setpoint = request.form.get("fed_batch_do_setpoint")
    induction_do_setpoint = request.form.get("induction_do_setpoint")
    target_batch_phase_airflow_vvm = request.form.get("target_batch_phase_airflow_vvm")
    target_fed_batch_phase_airflow_vvm = request.form.get("target_fed_batch_phase_airflow_vvm")
    foam_control = request.form.get("foam_control")
    inducer = request.form.get("inducer")
    target_batch_volume_ml = request.form.get("target_batch_volume_ml")
    target_pre_induction_volume_ml = request.form.get("target_pre_induction_volume_ml")
    target_final_volume_ml = request.form.get("target_final_volume_ml")
    target_final_volume_w_drawdown_ml = request.form.get("target_final_volume_w_drawdown_ml")
    seed_media = request.form.get("seed_media")
    feed_media_1 = request.form.get("feed_media_1")
    feed_media_1_density = request.form.get("feed_media_1_density")
    target_time_to_add_feed_media_1_post_feed_start_h = request.form.get("target_time_to_add_feed_media_1_post_feed_start_h")
    feed_media_2 = request.form.get("feed_media_2")
    target_time_to_add_feed_media_2_post_feed_start_h = request.form.get("target_time_to_add_feed_media_2_post_feed_start_h")
    estimated_batch_timing_h = request.form.get("estimated_batch_timing_h")
    target_post_feed_duration_until_temp_shift_h = request.form.get("target_post_feed_duration_until_temp_shift_h")
    target_post_feed_duration_until_induction_h = request.form.get("target_post_feed_duration_until_induction_h")
    target_induction_duration_h = request.form.get("target_induction_duration_h")
    growth_feed_rate_1_m_h_1 = request.form.get("growth_feed_rate_1_m_h_1")
    growth_feed_rate_1_post_feed_target_time_h = request.form.get("growth_feed_rate_1_post_feed_target_time_h")
    growth_feed_rate_1_type = request.form.get("growth_feed_rate_1_type")
    growth_feed_rate_2_m_h_1 = request.form.get("growth_feed_rate_2_m_h_1")
    growth_feed_rate_2_post_feed_target_time_h = request.form.get("growth_feed_rate_2_post_feed_target_time_h")
    growth_feed_rate_2_type = request.form.get("growth_feed_rate_2_type")
    induction_feed_rate_m_h_1 = request.form.get("induction_feed_rate_m_h_1")
    induction_feed_profile_type = request.form.get("induction_feed_profile_type")
    post_feed_time_of_percent_drawdown_h = request.form.get("post_feed_time_of_percent_drawdown_h")
    drawdown_at_post_feed_time_percent = request.form.get("drawdown_at_post_feed_time_percent")
    initial_batch_vol_litter = request.form.get("initial_batch_vol_litter")
    x0_grams_over_litter = request.form.get("x0_grams_over_litter")
    sf_feed_gluc_grams_over_litter = request.form.get("sf_feed_gluc_grams_over_litter")
    s0_batch_gluc_grams_over_litter = request.form.get("s0_batch_gluc_grams_over_litter")
    yx_over_s_grams_over_grams = request.form.get("yx_over_s_grams_over_grams")
    od_over_gram_dry_wt_over_litter = request.form.get("od_over_gram_dry_wt_over_litter")    
    growth_phase_feed_interval_h = request.form.get("growth_phase_feed_interval_h")
    feed_profile_type = request.form.get("feed_profile_type")
    induction_phase_feed_interval_h = request.form.get("induction_phase_feed_interval_h")
    actual_inoculation_time = request.form.get("actual_inoculation_time")
    actual_fed_batch_start_time = request.form.get("actual_fed_batch_start_time")
    actual_induction_start_time = request.form.get("actual_induction_start_time")
    target_induction_duration_1_h = request.form.get("target_induction_duration_1_h")
    addition_on_scale_pump_1 = request.form.get("addition_on_scale_pump_1")
    addition_on_scale_pump_2 = request.form.get("addition_on_scale_pump_2")
    if scale_liter == "":
        scale_liter = None
    if number_of_seed_stages == "":
        number_of_seed_stages = None
    if batch_temperature_setpoint_celcius == "":
        batch_temperature_setpoint_celcius = None
    if fed_batch_temperature_setpoint_celcius == "":
        fed_batch_temperature_setpoint_celcius = None
    if induction_temperature_setpoint_celcius == "":
        induction_temperature_setpoint_celcius = None
    if batch_phase_ph_setpoint == "":
        batch_phase_ph_setpoint = None
    if fed_batch_ph_setpoint == "":
        fed_batch_ph_setpoint = None
    if induction_ph_setpoint == "":
        induction_ph_setpoint = None
    if batch_phase_do_setpoint == "":
        batch_phase_do_setpoint = None
    if fed_batch_do_setpoint == "":
        fed_batch_do_setpoint = None
    if induction_do_setpoint == "":
        induction_do_setpoint = None
    if target_batch_phase_airflow_vvm == "":
        target_batch_phase_airflow_vvm = None
    if target_fed_batch_phase_airflow_vvm == "":
        target_fed_batch_phase_airflow_vvm = None
    if target_batch_volume_ml == "":
        target_batch_volume_ml = None
    if feed_media_1_density == "":
        feed_media_1_density = None
    if target_pre_induction_volume_ml == "":
        target_pre_induction_volume_ml = None
    if target_final_volume_ml == "":
        target_final_volume_ml = None
    if target_final_volume_w_drawdown_ml == "":
        target_final_volume_w_drawdown_ml = None
    if target_time_to_add_feed_media_2_post_feed_start_h == "":
        target_time_to_add_feed_media_2_post_feed_start_h = None
    if estimated_batch_timing_h == "":
        estimated_batch_timing_h = None
    if target_post_feed_duration_until_temp_shift_h == "":
        target_post_feed_duration_until_temp_shift_h = None
    if target_post_feed_duration_until_induction_h == "":
        target_post_feed_duration_until_induction_h = None
    if target_induction_duration_h == "":
        target_induction_duration_h = None
    if growth_feed_rate_1_m_h_1 == "":
        growth_feed_rate_1_m_h_1 = None
    if growth_feed_rate_1_post_feed_target_time_h == "":
        growth_feed_rate_1_post_feed_target_time_h = None
    if induction_feed_rate_m_h_1 == "":
        induction_feed_rate_m_h_1 = None
    if induction_phase_feed_interval_h == "":
        induction_phase_feed_interval_h = None
    if post_feed_time_of_percent_drawdown_h == "":
        post_feed_time_of_percent_drawdown_h = None
    if drawdown_at_post_feed_time_percent == "":
        drawdown_at_post_feed_time_percent = None
    if initial_batch_vol_litter == "":
        initial_batch_vol_litter = None
    if x0_grams_over_litter == "":
        x0_grams_over_litter = None
    if sf_feed_gluc_grams_over_litter == "":
        sf_feed_gluc_grams_over_litter = None
    if s0_batch_gluc_grams_over_litter == "":
        s0_batch_gluc_grams_over_litter = None
    if yx_over_s_grams_over_grams == "":
        yx_over_s_grams_over_grams = None
    if od_over_gram_dry_wt_over_litter == "":
        od_over_gram_dry_wt_over_litter = None
    if growth_phase_feed_interval_h == "":
        growth_phase_feed_interval_h = None
    if actual_inoculation_time == "":
        actual_inoculation_time = None
    if actual_fed_batch_start_time == "":
        actual_fed_batch_start_time = None
    if actual_induction_start_time == "":
        actual_induction_start_time = None
    if target_induction_duration_1_h == "":
        target_induction_duration_1_h = None
    if addition_on_scale_pump_1 == "":
        addition_on_scale_pump_1 = None
    if addition_on_scale_pump_2 == "":
        addition_on_scale_pump_2 = None
    try:
        trans = connection.begin()
        s = "DELETE FROM default_batch_process WHERE process_version_name_used_as_template='{}';".format(process_version_name_used_as_template)
        connection.execute(text(s))
        connection.execute(text("INSERT INTO pd_upstream.default_batch_process (process_version_name_used_as_template,process_description,strain,scale_liter,number_of_seed_stages,\
                    feed_medium_concentration,batch_temperature_setpoint_celcius,fed_batch_temperature_setpoint_celcius,induction_temperature_setpoint_celcius,\
                    batch_phase_ph_setpoint,fed_batch_ph_setpoint,induction_ph_setpoint,batch_phase_do_setpoint,fed_batch_do_setpoint,induction_do_setpoint,target_batch_phase_airflow_vvm,target_fed_batch_phase_airflow_vvm,foam_control,inducer,\
                    target_pre_induction_volume_ml,target_final_volume_ml,target_final_volume_w_drawdown_ml,seed_media,feed_media_1,feed_media_1_density,\
                    target_time_to_add_feed_media_1_post_feed_start_h,feed_media_2,target_time_to_add_feed_media_2_post_feed_start_h,estimated_batch_timing_h,\
                    target_post_feed_duration_until_temp_shift_h,target_post_feed_duration_until_induction_h,target_induction_duration_h,growth_feed_rate_1_m_h_1,\
                    growth_feed_rate_1_post_feed_target_time_h,growth_feed_rate_1_type,growth_feed_rate_2_m_h_1,growth_feed_rate_2_post_feed_target_time_h,\
                    growth_feed_rate_2_type,induction_feed_rate_m_h_1,induction_feed_profile_type,post_feed_time_of_percent_drawdown_h,drawdown_at_post_feed_time_percent,\
                    initial_batch_vol_litter,x0_grams_over_litter,sf_feed_gluc_grams_over_litter,s0_batch_gluc_grams_over_litter,yx_over_s_grams_over_grams,\
                    od_over_gram_dry_wt_over_litter,growth_phase_feed_interval_h,feed_profile_type,induction_phase_feed_interval_h,actual_inoculation_time,\
                    actual_fed_batch_start_time,actual_induction_start_time,target_induction_duration_1_h,target_batch_volume_ml,addition_on_scale_pump_1,addition_on_scale_pump_2)\
                    VALUES (:process_version_name_used_as_template,:process_description,:strain,:scale_liter,:number_of_seed_stages,\
                    :feed_medium_concentration,:batch_temperature_setpoint_celcius,:fed_batch_temperature_setpoint_celcius,:induction_temperature_setpoint_celcius,\
                    :batch_phase_ph_setpoint,:fed_batch_ph_setpoint,:induction_ph_setpoint,:batch_phase_do_setpoint,:fed_batch_do_setpoint,:induction_do_setpoint,:target_batch_phase_airflow_vvm,:target_fed_batch_phase_airflow_vvm,:foam_control,:inducer,\
                    :target_pre_induction_volume_ml,:target_final_volume_ml,:target_final_volume_w_drawdown_ml,:seed_media,:feed_media_1,:feed_media_1_density,\
                    :target_time_to_add_feed_media_1_post_feed_start_h,:feed_media_2,:target_time_to_add_feed_media_2_post_feed_start_h,:estimated_batch_timing_h,\
                    :target_post_feed_duration_until_temp_shift_h,:target_post_feed_duration_until_induction_h,:target_induction_duration_h,:growth_feed_rate_1_m_h_1,\
                    :growth_feed_rate_1_post_feed_target_time_h,:growth_feed_rate_1_type,:growth_feed_rate_2_m_h_1,:growth_feed_rate_2_post_feed_target_time_h,\
                    :growth_feed_rate_2_type,:induction_feed_rate_m_h_1,:induction_feed_profile_type,:post_feed_time_of_percent_drawdown_h,:drawdown_at_post_feed_time_percent,\
                    :initial_batch_vol_litter,:x0_grams_over_litter,:sf_feed_gluc_grams_over_litter,:s0_batch_gluc_grams_over_litter,:yx_over_s_grams_over_grams,\
                    :od_over_gram_dry_wt_over_litter,:growth_phase_feed_interval_h,:feed_profile_type,:induction_phase_feed_interval_h,:actual_inoculation_time,\
                    :actual_fed_batch_start_time,:actual_induction_start_time,:target_induction_duration_1_h,:target_batch_volume_ml,:addition_on_scale_pump_1,:addition_on_scale_pump_2)"),
                {"addition_on_scale_pump_1":addition_on_scale_pump_1,"addition_on_scale_pump_2":addition_on_scale_pump_2,"feed_media_1_density":feed_media_1_density,"process_version_name_used_as_template":process_version_name_used_as_template,"process_description":process_description,"strain":strain,"scale_liter":scale_liter,"number_of_seed_stages":number_of_seed_stages,"feed_medium_concentration":feed_medium_concentration,"batch_temperature_setpoint_celcius":batch_temperature_setpoint_celcius,"fed_batch_temperature_setpoint_celcius":fed_batch_temperature_setpoint_celcius,"induction_temperature_setpoint_celcius":induction_temperature_setpoint_celcius,"batch_phase_ph_setpoint":batch_phase_ph_setpoint,"fed_batch_ph_setpoint":fed_batch_ph_setpoint,"induction_ph_setpoint":induction_ph_setpoint,"batch_phase_do_setpoint":batch_phase_do_setpoint,"fed_batch_do_setpoint":fed_batch_do_setpoint,"induction_do_setpoint":induction_do_setpoint,"target_batch_phase_airflow_vvm":target_batch_phase_airflow_vvm,"target_fed_batch_phase_airflow_vvm":target_fed_batch_phase_airflow_vvm,"foam_control":foam_control,"inducer":inducer,"target_pre_induction_volume_ml":target_pre_induction_volume_ml,"target_final_volume_ml":target_final_volume_ml,"target_final_volume_w_drawdown_ml":target_final_volume_w_drawdown_ml,"seed_media":seed_media,"feed_media_1":feed_media_1,"target_time_to_add_feed_media_1_post_feed_start_h":target_time_to_add_feed_media_1_post_feed_start_h,"feed_media_2":feed_media_2,"target_time_to_add_feed_media_2_post_feed_start_h":target_time_to_add_feed_media_2_post_feed_start_h,"estimated_batch_timing_h":estimated_batch_timing_h,"target_post_feed_duration_until_temp_shift_h":target_post_feed_duration_until_temp_shift_h,"target_post_feed_duration_until_induction_h":target_post_feed_duration_until_induction_h,"target_induction_duration_h":target_induction_duration_h,"growth_feed_rate_1_m_h_1":growth_feed_rate_1_m_h_1,"growth_feed_rate_1_post_feed_target_time_h":growth_feed_rate_1_post_feed_target_time_h,"growth_feed_rate_1_type":growth_feed_rate_1_type,"growth_feed_rate_2_m_h_1":growth_feed_rate_2_m_h_1,"growth_feed_rate_2_post_feed_target_time_h":growth_feed_rate_2_post_feed_target_time_h,"growth_feed_rate_2_type":growth_feed_rate_2_type,"induction_feed_rate_m_h_1":induction_feed_rate_m_h_1,"induction_feed_profile_type":induction_feed_profile_type,"post_feed_time_of_percent_drawdown_h":post_feed_time_of_percent_drawdown_h,"drawdown_at_post_feed_time_percent":drawdown_at_post_feed_time_percent,"initial_batch_vol_litter":initial_batch_vol_litter,"x0_grams_over_litter":x0_grams_over_litter,"sf_feed_gluc_grams_over_litter":sf_feed_gluc_grams_over_litter,"s0_batch_gluc_grams_over_litter":s0_batch_gluc_grams_over_litter,"yx_over_s_grams_over_grams":yx_over_s_grams_over_grams,"od_over_gram_dry_wt_over_litter":od_over_gram_dry_wt_over_litter,"growth_phase_feed_interval_h":growth_phase_feed_interval_h,"feed_profile_type":feed_profile_type,"induction_phase_feed_interval_h":induction_phase_feed_interval_h,"actual_inoculation_time":actual_inoculation_time,"actual_fed_batch_start_time":actual_fed_batch_start_time,"actual_induction_start_time":actual_induction_start_time,"target_induction_duration_1_h":target_induction_duration_1_h,"target_batch_volume_ml":target_batch_volume_ml}) 
        connection.commit()
    except Exception:
        trans.rollback()
        raise
    return render_template("submitted.html")

@app.route('/batch_process_add')
def batch_process_add():
    try:
        #begin transaction
        trans=connection.begin()
        data = connection.execute(text("SELECT * FROM pd_upstream.default_batch_process"))
        insp = inspect(engine)
        tbl_cols = insp.get_columns('default_batch_process','pd_upstream')
        col_names = []
        for c in tbl_cols:
            col_names.append(c['name'])
        rows = []
        for r in data:
            rows.append(dict(zip(col_names, r)))

        processes = connection.execute(text("SELECT process_version_name_used_as_template FROM pd_upstream.default_batch_process order by process_version_name_used_as_template"))
        versions = []
        for v in processes:
            versions.append(v[0])
        trans.commit()

    except Exception:
        trans.rollback()
        raise
    return render_template("batch_process_add.html", rows=rows, versions=versions)

@app.route('/batch_process_add_submitted', methods=["POST"])
def batch_process_add_submitted():
    process_version_name_used_as_template = request.form.get("process_version_name_used_as_template")
    run_name = ''
    reactor_id = ''
    process_description = request.form.get("process_description")
    product = ''
    strain = request.form.get("strain")
    site_of_run = ''
    scale_liter = request.form.get("scale_liter")
    number_of_seed_stages = request.form.get("number_of_seed_stages")
    seed_train_run_id = ''
    feed_medium_concentration = request.form.get("feed_medium_concentration")
    batch_temperature_setpoint_celcius = request.form.get("batch_temperature_setpoint_celcius")
    fed_batch_temperature_setpoint_celcius = request.form.get("fed_batch_temperature_setpoint_celcius")
    induction_temperature_setpoint_celcius = request.form.get("induction_temperature_setpoint_celcius")
    batch_phase_ph_setpoint = request.form.get("batch_phase_ph_setpoint")
    fed_batch_ph_setpoint = request.form.get("fed_batch_ph_setpoint")
    induction_ph_setpoint = request.form.get("induction_ph_setpoint")
    batch_phase_do_setpoint = request.form.get("batch_phase_do_setpoint")
    fed_batch_do_setpoint = request.form.get("fed_batch_do_setpoint")
    induction_do_setpoint = request.form.get("induction_do_setpoint")
    target_batch_phase_airflow_vvm = request.form.get("target_batch_phase_airflow_vvm")
    target_fed_batch_phase_airflow_vvm = request.form.get("target_fed_batch_phase_airflow_vvm")
    foam_control = request.form.get("foam_control")
    inducer = request.form.get("inducer")
    target_batch_volume_ml = request.form.get("target_batch_volume_ml")
    target_pre_induction_volume_ml = request.form.get("target_pre_induction_volume_ml")
    target_final_volume_ml = request.form.get("target_final_volume_ml")
    target_final_volume_w_drawdown_ml = request.form.get("target_final_volume_w_drawdown_ml")
    seed_media = request.form.get("seed_media")
    feed_media_1 = request.form.get("feed_media_1")
    feed_media_1_density = request.form.get("feed_media_1_density")
    target_time_to_add_feed_media_1_post_feed_start_h = request.form.get("target_time_to_add_feed_media_1_post_feed_start_h")
    feed_media_2 = request.form.get("feed_media_2")
    target_time_to_add_feed_media_2_post_feed_start_h = request.form.get("target_time_to_add_feed_media_2_post_feed_start_h")
    estimated_batch_timing_h = request.form.get("estimated_batch_timing_h")
    target_post_feed_duration_until_temp_shift_h = request.form.get("target_post_feed_duration_until_temp_shift_h")
    target_post_feed_duration_until_induction_h = request.form.get("target_post_feed_duration_until_induction_h")
    target_induction_duration_h = request.form.get("target_induction_duration_h")
    growth_feed_rate_1_m_h_1 = request.form.get("growth_feed_rate_1_m_h_1")
    growth_feed_rate_1_post_feed_target_time_h = request.form.get("growth_feed_rate_1_post_feed_target_time_h")
    growth_feed_rate_1_type = request.form.get("growth_feed_rate_1_type")
    growth_feed_rate_2_m_h_1 = request.form.get("growth_feed_rate_2_m_h_1")
    growth_feed_rate_2_post_feed_target_time_h = request.form.get("growth_feed_rate_2_post_feed_target_time_h")
    growth_feed_rate_2_type = request.form.get("growth_feed_rate_2_type")
    induction_feed_rate_m_h_1 = request.form.get("induction_feed_rate_m_h_1")
    induction_feed_profile_type = request.form.get("induction_feed_profile_type")
    post_feed_time_of_percent_drawdown_h = request.form.get("post_feed_time_of_percent_drawdown_h")
    drawdown_at_post_feed_time_percent = request.form.get("drawdown_at_post_feed_time_percent")
    initial_batch_vol_litter = request.form.get("initial_batch_vol_litter")
    x0_grams_over_litter = request.form.get("x0_grams_over_litter")
    sf_feed_gluc_grams_over_litter = request.form.get("sf_feed_gluc_grams_over_litter")
    s0_batch_gluc_grams_over_litter = request.form.get("s0_batch_gluc_grams_over_litter")
    yx_over_s_grams_over_grams = request.form.get("yx_over_s_grams_over_grams")
    od_over_gram_dry_wt_over_litter = request.form.get("od_over_gram_dry_wt_over_litter")    
    growth_phase_feed_interval_h = request.form.get("growth_phase_feed_interval_h")
    feed_profile_type = request.form.get("feed_profile_type")
    induction_phase_feed_interval_h = request.form.get("induction_phase_feed_interval_h")
    actual_inoculation_time = request.form.get("actual_inoculation_time")
    actual_fed_batch_start_time = request.form.get("actual_fed_batch_start_time")
    actual_induction_start_time = request.form.get("actual_induction_start_time")
    target_induction_duration_1_h = request.form.get("target_induction_duration_1_h")
    addition_on_scale_pump_1 = request.form.get("addition_on_scale_pump_1")
    addition_on_scale_pump_2 = request.form.get("addition_on_scale_pump_2")
    run_outcome = ''
    if addition_on_scale_pump_1 == "":
        addition_on_scale_pump_1 = None
    if addition_on_scale_pump_2 == "":
        addition_on_scale_pump_2 = None
    if scale_liter == "":
        scale_liter = None
    if number_of_seed_stages == "":
        number_of_seed_stages = None
    if batch_temperature_setpoint_celcius == "":
        batch_temperature_setpoint_celcius = None
    if fed_batch_temperature_setpoint_celcius == "":
        fed_batch_temperature_setpoint_celcius = None
    if induction_temperature_setpoint_celcius == "":
        induction_temperature_setpoint_celcius = None
    if batch_phase_ph_setpoint == "":
        batch_phase_ph_setpoint = None
    if fed_batch_ph_setpoint == "":
        fed_batch_ph_setpoint = None
    if feed_media_1_density == "":
        feed_media_1_density = None
    if induction_ph_setpoint == "":
        induction_ph_setpoint = None
    if batch_phase_do_setpoint == "":
        batch_phase_do_setpoint = None
    if fed_batch_do_setpoint == "":
        fed_batch_do_setpoint = None
    if induction_do_setpoint == "":
        induction_do_setpoint = None
    if target_batch_phase_airflow_vvm == "":
        target_batch_phase_airflow_vvm = None
    if target_fed_batch_phase_airflow_vvm == "":
        target_fed_batch_phase_airflow_vvm = None
    if target_batch_volume_ml == "":
        target_batch_volume_ml = None
    if target_pre_induction_volume_ml == "":
        target_pre_induction_volume_ml = None
    if target_final_volume_ml == "":
        target_final_volume_ml = None
    if target_final_volume_w_drawdown_ml == "":
        target_final_volume_w_drawdown_ml = None
    if target_time_to_add_feed_media_2_post_feed_start_h == "":
        target_time_to_add_feed_media_2_post_feed_start_h = None
    if estimated_batch_timing_h == "":
        estimated_batch_timing_h = None
    if target_post_feed_duration_until_temp_shift_h == "":
        target_post_feed_duration_until_temp_shift_h = None
    if target_post_feed_duration_until_induction_h == "":
        target_post_feed_duration_until_induction_h = None
    if target_induction_duration_h == "":
        target_induction_duration_h = None
    if growth_feed_rate_1_m_h_1 == "":
        growth_feed_rate_1_m_h_1 = None
    if growth_feed_rate_1_post_feed_target_time_h == "":
        growth_feed_rate_1_post_feed_target_time_h = None
    if induction_feed_rate_m_h_1 == "":
        induction_feed_rate_m_h_1 = None
    if induction_phase_feed_interval_h == "":
        induction_phase_feed_interval_h = None
    if post_feed_time_of_percent_drawdown_h == "":
        post_feed_time_of_percent_drawdown_h = None
    if drawdown_at_post_feed_time_percent == "":
        drawdown_at_post_feed_time_percent = None
    if initial_batch_vol_litter == "":
        initial_batch_vol_litter = None
    if x0_grams_over_litter == "":
        x0_grams_over_litter = None
    if sf_feed_gluc_grams_over_litter == "":
        sf_feed_gluc_grams_over_litter = None
    if s0_batch_gluc_grams_over_litter == "":
        s0_batch_gluc_grams_over_litter = None
    if yx_over_s_grams_over_grams == "":
        yx_over_s_grams_over_grams = None
    if od_over_gram_dry_wt_over_litter == "":
        od_over_gram_dry_wt_over_litter = None
    if growth_phase_feed_interval_h == "":
        growth_phase_feed_interval_h = None
    if actual_inoculation_time == "":
        actual_inoculation_time = None
    if actual_fed_batch_start_time == "":
        actual_fed_batch_start_time = None
    if actual_induction_start_time == "":
        actual_induction_start_time = None
    if target_induction_duration_1_h == "":
        target_induction_duration_1_h = None
    try:
        trans = connection.begin()
        connection.execute(text("INSERT INTO pd_upstream.default_batch_process (process_version_name_used_as_template,run_name,reactor_id,process_description,product,strain,site_of_run,scale_liter,number_of_seed_stages,seed_train_run_id,\
                    feed_medium_concentration,batch_temperature_setpoint_celcius,fed_batch_temperature_setpoint_celcius,induction_temperature_setpoint_celcius,\
                    batch_phase_ph_setpoint,fed_batch_ph_setpoint,induction_ph_setpoint,batch_phase_do_setpoint,fed_batch_do_setpoint,induction_do_setpoint,target_batch_phase_airflow_vvm,target_fed_batch_phase_airflow_vvm,foam_control,inducer,\
                    target_pre_induction_volume_ml,target_final_volume_ml,target_final_volume_w_drawdown_ml,seed_media,feed_media_1,feed_media_1_density,\
                    target_time_to_add_feed_media_1_post_feed_start_h,feed_media_2,target_time_to_add_feed_media_2_post_feed_start_h,estimated_batch_timing_h,\
                    target_post_feed_duration_until_temp_shift_h,target_post_feed_duration_until_induction_h,target_induction_duration_h,growth_feed_rate_1_m_h_1,\
                    growth_feed_rate_1_post_feed_target_time_h,growth_feed_rate_1_type,growth_feed_rate_2_m_h_1,growth_feed_rate_2_post_feed_target_time_h,\
                    growth_feed_rate_2_type,induction_feed_rate_m_h_1,induction_feed_profile_type,post_feed_time_of_percent_drawdown_h,drawdown_at_post_feed_time_percent,\
                    initial_batch_vol_litter,x0_grams_over_litter,sf_feed_gluc_grams_over_litter,s0_batch_gluc_grams_over_litter,yx_over_s_grams_over_grams,\
                    od_over_gram_dry_wt_over_litter,growth_phase_feed_interval_h,feed_profile_type,induction_phase_feed_interval_h,actual_inoculation_time,\
                    actual_fed_batch_start_time,actual_induction_start_time,target_induction_duration_1_h,run_outcome,addition_on_scale_pump_1,addition_on_scale_pump_2)\
                    VALUES (:process_version_name_used_as_template,:run_name,:reactor_id,:process_description,:product,:strain,:site_of_run,:scale_liter,:number_of_seed_stages,:seed_train_run_id,\
                    :feed_medium_concentration,:batch_temperature_setpoint_celcius,:fed_batch_temperature_setpoint_celcius,:induction_temperature_setpoint_celcius,\
                    :batch_phase_ph_setpoint,:fed_batch_ph_setpoint,:induction_ph_setpoint,:batch_phase_do_setpoint,:fed_batch_do_setpoint,:induction_do_setpoint,:target_batch_phase_airflow_vvm,:target_fed_batch_phase_airflow_vvm,:foam_control,:inducer,\
                    :target_pre_induction_volume_ml,:target_final_volume_ml,:target_final_volume_w_drawdown_ml,:seed_media,:feed_media_1,:feed_media_1_density,\
                    :target_time_to_add_feed_media_1_post_feed_start_h,:feed_media_2,:target_time_to_add_feed_media_2_post_feed_start_h,:estimated_batch_timing_h,\
                    :target_post_feed_duration_until_temp_shift_h,:target_post_feed_duration_until_induction_h,:target_induction_duration_h,:growth_feed_rate_1_m_h_1,\
                    :growth_feed_rate_1_post_feed_target_time_h,:growth_feed_rate_1_type,:growth_feed_rate_2_m_h_1,:growth_feed_rate_2_post_feed_target_time_h,\
                    :growth_feed_rate_2_type,:induction_feed_rate_m_h_1,:induction_feed_profile_type,:post_feed_time_of_percent_drawdown_h,:drawdown_at_post_feed_time_percent,\
                    :initial_batch_vol_litter,:x0_grams_over_litter,:sf_feed_gluc_grams_over_litter,:s0_batch_gluc_grams_over_litter,:yx_over_s_grams_over_grams,\
                    :od_over_gram_dry_wt_over_litter,:growth_phase_feed_interval_h,:feed_profile_type,:induction_phase_feed_interval_h,:actual_inoculation_time,\
                    :actual_fed_batch_start_time,:actual_induction_start_time,:target_induction_duration_1_h,:run_outcome,:addition_on_scale_pump_1,:addition_on_scale_pump_2)"),
                {"addition_on_scale_pump_1":addition_on_scale_pump_1,"addition_on_scale_pump_2":addition_on_scale_pump_2,"feed_media_1_density":feed_media_1_density,"process_version_name_used_as_template":process_version_name_used_as_template,"run_name":run_name,"reactor_id":reactor_id,"process_description":process_description,"product":product,"strain":strain,"site_of_run":site_of_run,"scale_liter":scale_liter,"number_of_seed_stages":number_of_seed_stages,"seed_train_run_id":seed_train_run_id,"feed_medium_concentration":feed_medium_concentration,"batch_temperature_setpoint_celcius":batch_temperature_setpoint_celcius,"fed_batch_temperature_setpoint_celcius":fed_batch_temperature_setpoint_celcius,"induction_temperature_setpoint_celcius":induction_temperature_setpoint_celcius,"batch_phase_ph_setpoint":batch_phase_ph_setpoint,"fed_batch_ph_setpoint":fed_batch_ph_setpoint,"induction_ph_setpoint":induction_ph_setpoint,"batch_phase_do_setpoint":batch_phase_do_setpoint,"fed_batch_do_setpoint":fed_batch_do_setpoint,"induction_do_setpoint":induction_do_setpoint,"target_batch_phase_airflow_vvm":target_batch_phase_airflow_vvm,"target_fed_batch_phase_airflow_vvm":target_fed_batch_phase_airflow_vvm,"foam_control":foam_control,"inducer":inducer,"target_pre_induction_volume_ml":target_pre_induction_volume_ml,"target_final_volume_ml":target_final_volume_ml,"target_final_volume_w_drawdown_ml":target_final_volume_w_drawdown_ml,"seed_media":seed_media,"feed_media_1":feed_media_1,"target_time_to_add_feed_media_1_post_feed_start_h":target_time_to_add_feed_media_1_post_feed_start_h,"feed_media_2":feed_media_2,"target_time_to_add_feed_media_2_post_feed_start_h":target_time_to_add_feed_media_2_post_feed_start_h,"estimated_batch_timing_h":estimated_batch_timing_h,"target_post_feed_duration_until_temp_shift_h":target_post_feed_duration_until_temp_shift_h,"target_post_feed_duration_until_induction_h":target_post_feed_duration_until_induction_h,"target_induction_duration_h":target_induction_duration_h,"growth_feed_rate_1_m_h_1":growth_feed_rate_1_m_h_1,"growth_feed_rate_1_post_feed_target_time_h":growth_feed_rate_1_post_feed_target_time_h,"growth_feed_rate_1_type":growth_feed_rate_1_type,"growth_feed_rate_2_m_h_1":growth_feed_rate_2_m_h_1,"growth_feed_rate_2_post_feed_target_time_h":growth_feed_rate_2_post_feed_target_time_h,"growth_feed_rate_2_type":growth_feed_rate_2_type,"induction_feed_rate_m_h_1":induction_feed_rate_m_h_1,"induction_feed_profile_type":induction_feed_profile_type,"post_feed_time_of_percent_drawdown_h":post_feed_time_of_percent_drawdown_h,"drawdown_at_post_feed_time_percent":drawdown_at_post_feed_time_percent,"initial_batch_vol_litter":initial_batch_vol_litter,"x0_grams_over_litter":x0_grams_over_litter,"sf_feed_gluc_grams_over_litter":sf_feed_gluc_grams_over_litter,"s0_batch_gluc_grams_over_litter":s0_batch_gluc_grams_over_litter,"yx_over_s_grams_over_grams":yx_over_s_grams_over_grams,"od_over_gram_dry_wt_over_litter":od_over_gram_dry_wt_over_litter,"growth_phase_feed_interval_h":growth_phase_feed_interval_h,"feed_profile_type":feed_profile_type,"induction_phase_feed_interval_h":induction_phase_feed_interval_h,"actual_inoculation_time":actual_inoculation_time,"actual_fed_batch_start_time":actual_fed_batch_start_time,"actual_induction_start_time":actual_induction_start_time,"target_induction_duration_1_h":target_induction_duration_1_h,"run_outcome":run_outcome}) 
        connection.commit()
    except Exception:
        trans.rollback()
        raise
    return render_template("submitted.html")

@app.route('/addition_process_edit')
def addition_process_edit():
    try:
        trans = connection.begin()
        data = connection.execute("SELECT * FROM pd_upstream.default_addition_process")
        insp = inspect(engine)
        tbl_cols = insp.get_columns('default_addition_process','pd_upstream')
        col_names = []
        for c in tbl_cols:
            col_names.append(c['name'])
        rows = []
        for r in data:
            rows.append(dict(zip(col_names, r)))
        processes = connection.execute(text("SELECT process_version_name_used_as_template FROM pd_upstream.default_addition_process order by process_version_name_used_as_template"))

        addition_versions = []
        for v in processes:
            if not v[0] in addition_versions:
                addition_versions.append(v[0])
        
        phases = ['Pre-Sterilization','Post-Sterilization','Inoculation','Batch Phase','Fed Batch Growth Phase','Fed Batch Induction Phase']
        trans.commit()
    except Exception:
        trans.rollback()
        raise
    return render_template("addition_process_edit.html",addition_versions=addition_versions,rows=rows,phases=phases)

@app.route('/addition_process_edit_submitted', methods=["POST"])
def addition_process_edit_submitted():
    process_version_name_used_as_template = request.form.get('process_version_name_used_as_template')
    addition_name = request.form.get('addition_name')
    phase = request.form.get('phase')
    target_phase_time_h = request.form.get('target_phase_time_h')
    target_batch_concentration_ml_li = request.form.get('target_batch_concentration_ml_li')
    if target_batch_concentration_ml_li == "":
        target_batch_concentration_ml_li = None
    try:
        trans = connection.begin()

        s = "DELETE FROM default_addition_process WHERE process_version_name_used_as_template='{}';".format(process_version_name_used_as_template)
        connection.execute(text(s))
        connection.execute(text("INSERT INTO pd_upstream.default_addition_process (process_version_name_used_as_template,addition_name,phase,target_phase_time_h,target_batch_concentration_ml_li) \
                    VALUES (:process_version_name_used_as_template,:addition_name,:phase,:target_phase_time_h,:target_batch_concentration_ml_li)"),
            {"process_version_name_used_as_template":process_version_name_used_as_template,"addition_name":addition_name,"phase":phase,"target_phase_time_h":target_phase_time_h,"target_batch_concentration_ml_li":target_batch_concentration_ml_li}) 
        connection.commit()
    except Exception:
        trans.rollback()
        raise
    
    return render_template("submitted.html")

@app.route('/addition_process_add')
def addition_process_add():
    form = AdditionPrcoessVersion()
    phases = ['Pre-Sterilization','Post-Sterilization','Inoculation','Batch Phase','Fed Batch Growth Phase','Fed Batch Induction Phase']
    try:
        trans = connection.begin()
        processes = connection.execute(text("SELECT process_version_name_used_as_template FROM pd_upstream.default_addition_process order by process_version_name_used_as_template"))
        addition_versions = []
        for v in processes:
            if not v[0] in addition_versions:
                addition_versions.append(v[0])
    except Exception:
        trans.rollback()
        raise
    return render_template("addition_process_add.html",form=form,phases=phases,addition_versions=addition_versions)

@app.route('/addition_process_add_submitted', methods=["POST"])
def addition_process_add_submitted():
    process_version_name_used_as_template = request.form.get('process_version_name_used_as_template')
    addition_name = request.form.get('addition_name')
    phase = request.form.get('phase')
    target_phase_time_h = request.form.get('target_phase_time_h')
    target_batch_concentration_ml_li = request.form.get('target_batch_concentration_ml_li')
    if target_batch_concentration_ml_li == "":
        target_batch_concentration_ml_li = None
    try:
        trans = connection.begin()
        connection.execute(text("INSERT INTO pd_upstream.default_addition_process (process_version_name_used_as_template,addition_name,phase,target_phase_time_h,target_batch_concentration_ml_li) \
                    VALUES (:process_version_name_used_as_template,:addition_name,:phase,:target_phase_time_h,:target_batch_concentration_ml_li)"),
            {"process_version_name_used_as_template":process_version_name_used_as_template,"addition_name":addition_name,"phase":phase,"target_phase_time_h":target_phase_time_h,"target_batch_concentration_ml_li":target_batch_concentration_ml_li}) 
        connection.commit()
    except Exception:
        trans.rollback()
        raise
    return render_template("submitted.html") 

@app.route('/summary')
def summary():
    #experiment_d = db.execute("SELECT * FROM pd_upstream.experiment")
    #experiment_data = [dict(experiment_data) for experiment_data in experiment_d.fetchall()]
    db.execute('SELECT * FROM pd_upstream.experiment')
    experiment_d = db.fetchall()
    db.execute("Select * FROM pd_upstream.experiment")
    #db.execute("Select * FROM default_batch_process LIMIT 0")
    exp_keys = [desc[0] for desc in db.description]
    experiment_data =[]
    for experiment in experiment_d:
        experiment_data.append(dict(zip(exp_keys,experiment)))
    

    #run_d = db.execute("SELECT * FROM pd_upstream.sutro_production_batch_record")
    #run_data = [dict(run_data) for run_data in run_d.fetchall()]
    db.execute('SELECT * FROM pd_upstream.sutro_production_batch_record')
    run_d = db.fetchall()
    db.execute("Select * FROM pd_upstream.sutro_production_batch_record")

    #db.execute("Select * FROM pd_upstream.sutro_production_batch_record LIMIT 0")
    run_keys = [desc[0] for desc in db.description]
    run_data =[]
    for run in run_d:
        run_data.append(dict(zip(run_keys,run)))

    #timepoint_d = db.execute("SELECT * FROM pd_upstream.growth_data")
    #timepoint_data = [dict(timepoint_data) for timepoint_data in timepoint_d.fetchall()]
    db.execute('SELECT * FROM pd_upstream.growth_data')
    timepoint_d=db.fetchall()
    db.execute('SELECT * FROM pd_upstream.growth_data')
    #db.execute("Select * FROM pd_upstream.growth_data LIMIT 0")
    timept_keys = [desc[0] for desc in db.description]
    timepoint_data =[]
    for tp in timepoint_d:
        timepoint_data.append(dict(zip(timept_keys,tp)))

    
    db.execute('SELECT * FROM pd_upstream.production_addition')
    addition_d=db.fetchall()
    db.execute('SELECT * FROM pd_upstream.production_addition')
    add_keys = [desc[0] for desc in db.description]
    addition_data = []
    for addition in addition_d:
        addition_data.append(dict(zip(add_keys,addition)))

    
    
    return render_template("summary.html",addition_data=addition_data,experiment_data=experiment_data,run_data=run_data,timepoint_data=timepoint_data)

@app.route('/online_bioreactor_data_upload')
def online_bioreactor_data_upload():
  

    db.execute("SELECT * FROM pd_upstream.mfcs_online_data_file_list")
    mfcs_file_list_rows =  db.fetchall()
    run_data = []
    for row in mfcs_file_list_rows:
        run_data.append(row[0])
    db.execute("SELECT run_name FROM pd_upstream.sutro_production_batch_record order by id desc")
    run_name_rows = db.fetchall()
    run_names_lst = []
    for r in run_name_rows:
        run_names_lst.append(r[0])

    return render_template('online_bioreactor_data_upload.html',run_data=run_data,run_names_lst=run_names_lst)

@app.route('/online_bioreactor_data_upload_submitted', methods=['GET','POST'])
def online_bioreactor_data_upload_submitted():
    db_name_to_excel_name = {
            "pdattime":"PDatTime",
            "age_h":"Age (h)",
            "temp_value":"TEMP_Value",
            "_batchalarmcount_value":"_BatchAlarmCount_Value",
            "_batchsync_value":"_BatchSync_Value",
            "ph_value":"pH_Value",
            "po2_value":"pO2_Value",
            "stirr_value":"STIRR_Value",
            "_batchage_value":"_BatchAge_Value",
            "acidt_value":"ACIDT_Value",
            "baset_value":"BASET_Value",
            "ext_value":"EXT_Value",
            "fo_le_t_value":"FO_LE_T_Value",
            "fo_le_t_setpoint":"FO_LE_T_Setpoint",
            "gasfl_value":"GASFL_Value",
            "gasfl_setpoint":"GASFL_Setpoint",
            "jtemp_value":"JTEMP_Value",
            "o2en_value":"O2EN_Value",
            "o2en_setpoint":"O2EN_Setpoint",
            "ph_setpoint":"pH_Setpoint",
            "po2_setpoint":"pO2_Setpoint",
            "stirr_setpoint":"STIRR_Setpoint",
            "subsa_value":"SUBSA_Value",
            "subsa_setpoint":"SUBSA_Setpoint",
            "subsb_value":"SUBSB_Value",
            "subsb_setpoint":"SUBSB_Setpoint",
            "temp_setpoint":"TEMP_Setpoint" 
        }
    num_runs = request.form.get('num_runs')
    # Each file will take about ~ 1.5 minutes to upload (for 65k rows in the Excel file), 3 minutes if run id already in mfcs_qs
    for run in range(int(num_runs)):
        run_id = request.form.get('run_id_' + str(run))
        #st = "SELECT id FROM pd_upstream.sutro_production_batch_record WHERE run_name='{}'".format(run_id)
        db.execute("SELECT id from pd_upstream.sutro_production_batch_record WHERE run_name = %(run_id)s", {"run_id":run_id})
        #actual_run_id_query = db.execute(st, run_id)
        actual_run_id_query = db.fetchall()
        actual_run_id = 0
        #for i in actual_run_id_query:
            #actual_run_id = int(i[0])
        for i in actual_run_id_query:
            actual_run_id=int(i[0])
        # associates with the most recently added row (to sutro production batch record) of this run id (but ideally the same run id shouldn't be in there more than once)
        
        #s = "DELETE FROM pd_upstream.mfcs_qs WHERE run_id='{}';".format(run_id)
        db.execute("DELETE FROM pd_upstream.mfcs_qs WHERE run_id_m= %(run_id)s", {"run_id":run_id})
        #db.execute(s)
        #db.commit()
        conn.commit()
        f = request.files['file_' + str(run)]
        filename = f.filename
        #q = "DELETE FROM pd_upstream.mfcs_online_data_file_list WHERE run_id='{}';".format(run_id)
        #db.execute(q)
        db.execute("DELETE FROM pd_upstream.mfcs_online_data_file_list WHERE run_id= %(run_id)s", {"run_id":run_id})
        #db.commit()
        conn.commit()
        #db.execute("INSERT INTO pd_upstream.mfcs_online_data_file_list (filename,run_id) VALUES (:filename,:run_id)",{"filename":filename,"run_id":run_id})
        #db.commit()
        db.execute("""INSERT INTO pd_upstream.mfcs_online_data_file_list (filename,run_id) VALUES (%(filename)s,%(run_id)s)""",{"filename":filename,"run_id":run_id})

        conn.commit()
        df = pd.read_excel(f)
        for i,row in df.iterrows():
            try:
                pdattime = row[db_name_to_excel_name['pdattime']]
            except KeyError:
                pdattime = None
            try:
                age_h = row[db_name_to_excel_name['age_h']]
            except KeyError:
                age_h = None
            try:
                temp_value = row[db_name_to_excel_name['temp_value']]
            except KeyError:
                temp_value = None
            try:
                _batchalarmcount_value = row[db_name_to_excel_name['_batchalarmcount_value']]
            except KeyError:
                _batchalarmcount_value = None
            try:
                _batchsync_value = row[db_name_to_excel_name['_batchsync_value']]
            except KeyError:
                _batchsync_value = None
            try:
                ph_value = row[db_name_to_excel_name['ph_value']]
            except KeyError:
                ph_value = None
            try:
                po2_value = row[db_name_to_excel_name['po2_value']]
            except KeyError:
                po2_value = None
            try:
                stirr_value = row[db_name_to_excel_name['stirr_value']]
            except KeyError:
                stirr_value = None
            try:
                _batchage_value = row[db_name_to_excel_name['_batchage_value']]
            except KeyError:
                _batchage_value = None
            try:
                acidt_value = row[db_name_to_excel_name['acidt_value']]
            except KeyError:
                acidt_value = None
            try:
                baset_value = row[db_name_to_excel_name['baset_value']]
            except KeyError:
                baset_value = None
            try:
                ext_value = row[db_name_to_excel_name['ext_value']]
            except KeyError:
                ext_value = None
            try:
                fo_le_t_value = row[db_name_to_excel_name['fo_le_t_value']]
            except KeyError:
                fo_le_t_value = None
            try:
                fo_le_t_setpoint = row[db_name_to_excel_name['fo_le_t_setpoint']]
            except KeyError:
                fo_le_t_setpoint = None
            try:
                gasfl_value = row[db_name_to_excel_name['gasfl_value']]
            except KeyError:
                gasfl_value = None
            try:
                gasfl_setpoint = row[db_name_to_excel_name['gasfl_setpoint']]
            except KeyError:
                gasfl_setpoint = None
            try:
                jtemp_value = row[db_name_to_excel_name['jtemp_value']]
            except KeyError:
                jtemp_value = None
            try:
                o2en_value = row[db_name_to_excel_name['o2en_value']]
            except KeyError:
                o2en_value = None
            try:
                o2en_setpoint = row[db_name_to_excel_name['o2en_setpoint']]
            except KeyError:
                o2en_setpoint = None
            try:
                ph_setpoint = row[db_name_to_excel_name['ph_setpoint']]
            except KeyError:
                ph_setpoint = None
            try:
                po2_setpoint = row[db_name_to_excel_name['po2_setpoint']]
            except KeyError:
                po2_setpoint = None
            try:
                stirr_setpoint = row[db_name_to_excel_name['stirr_setpoint']]
            except KeyError:
                stirr_setpoint = None
            try:
                subsa_value = row[db_name_to_excel_name['subsa_value']]
            except KeyError:
                subsa_value = None
            try:
                subsa_setpoint = row[db_name_to_excel_name['subsa_setpoint']]
            except KeyError:
                subsa_setpoint = None
            try:
                subsb_value = row[db_name_to_excel_name['subsb_value']]
            except KeyError:
                subsb_value = None
            try:
                subsb_setpoint = row[db_name_to_excel_name['subsb_setpoint']]
            except KeyError:
                subsb_setpoint = None
            try:
                temp_setpoint = row[db_name_to_excel_name['temp_setpoint']]
            except KeyError:
                temp_setpoint = None
            #db.execute("INSERT INTO pd_upstream.mfcs_qs (filename,pdattime,age_h,temp_value,stirr_value,ph_value,po2_value,_batchage_value,ph_setpoint,po2_setpoint,stirr_setpoint,temp_setpoint,_batchalarmcount_value,_batchsync_value,acidt_value,baset_value,ext_value,fo_le_t_value,fo_le_t_setpoint,gasfl_value,gasfl_setpoint,jtemp_value,o2en_value,o2en_setpoint,subsa_setpoint,subsb_setpoint,run_id,actual_run_id) \
                       # VALUES (:filename,:pdattime,:age_h,:temp_value,:stirr_value,:ph_value,:po2_value,:_batchage_value,:ph_setpoint,:po2_setpoint,:stirr_setpoint,:temp_setpoint,:_batchalarmcount_value,:_batchsync_value,:acidt_value,:baset_value,:ext_value,:fo_le_t_value,:fo_le_t_setpoint,:gasfl_value,:gasfl_setpoint,:jtemp_value,:o2en_value,:o2en_setpoint,:subsa_setpoint,:subsb_setpoint,:run_id,:actual_run_id)",
                #{"filename":filename,"pdattime":pdattime,"age_h":age_h,"temp_value":temp_value,"stirr_value":stirr_value,"ph_value":ph_value,"po2_value":po2_value,"_batchage_value":_batchage_value,"ph_setpoint":ph_setpoint,"po2_setpoint":po2_setpoint,"stirr_setpoint":stirr_setpoint,"temp_setpoint":temp_setpoint,"_batchalarmcount_value":_batchalarmcount_value,"_batchsync_value":_batchsync_value,"acidt_value":acidt_value,"baset_value":baset_value,"ext_value":ext_value,"fo_le_t_value":fo_le_t_value,"fo_le_t_setpoint":fo_le_t_setpoint,"gasfl_value":gasfl_value,"gasfl_setpoint":gasfl_setpoint,"jtemp_value":jtemp_value,"o2en_value":o2en_value,"o2en_setpoint":o2en_setpoint,"subsa_setpoint":subsa_setpoint,"subsb_setpoint":subsb_setpoint,"run_id":run_id,"actual_run_id":actual_run_id}) 
            #db.commit()
            db.execute("""INSERT INTO pd_upstream.mfcs_qs (filename,pdattime,age_h,temp_value,stirr_value,ph_value,po2_value,_batchage_value,ph_setpoint,po2_setpoint,stirr_setpoint,temp_setpoint,_batchalarmcount_value,_batchsync_value,acidt_value,baset_value,ext_value,fo_le_t_value,fo_le_t_setpoint,gasfl_value,gasfl_setpoint,jtemp_value,o2en_value,o2en_setpoint,subsa_setpoint,subsb_setpoint,run_id_m,actual_run_id) \
                    VALUES (%(filename)s,%(pdattime)s,%(age_h)s,%(temp_value)s,%(stirr_value)s,%(ph_value)s,%(po2_value)s,%(_batchage_value)s,%(ph_setpoint)s,%(po2_setpoint)s,%(stirr_setpoint)s,%(temp_setpoint)s,%(_batchalarmcount_value)s,%(_batchsync_value)s,%(acidt_value)s,%(baset_value)s,%(ext_value)s,%(fo_le_t_value)s,%(fo_le_t_setpoint)s,%(gasfl_value)s,%(gasfl_setpoint)s,%(jtemp_value)s,%(o2en_value)s,%(o2en_setpoint)s,%(subsa_setpoint)s,%(subsb_setpoint)s,%(run_id)s,%(actual_run_id)s)""",                 
                    {"filename":filename,"pdattime":pdattime,"age_h":age_h,"temp_value":temp_value,"stirr_value":stirr_value,"ph_value":ph_value,"po2_value":po2_value,"_batchage_value":_batchage_value,"ph_setpoint":ph_setpoint,"po2_setpoint":po2_setpoint,"stirr_setpoint":stirr_setpoint,"temp_setpoint":temp_setpoint,"_batchalarmcount_value":_batchalarmcount_value,"_batchsync_value":_batchsync_value,"acidt_value":acidt_value,"baset_value":baset_value,"ext_value":ext_value,"fo_le_t_value":fo_le_t_value,"fo_le_t_setpoint":fo_le_t_setpoint,"gasfl_value":gasfl_value,"gasfl_setpoint":gasfl_setpoint,"jtemp_value":jtemp_value,"o2en_value":o2en_value,"o2en_setpoint":o2en_setpoint,"subsa_setpoint":subsa_setpoint,"subsb_setpoint":subsb_setpoint,"run_id":run_id,"actual_run_id":actual_run_id}) 
                


            conn.commit()
    return render_template('submitted.html')

@app.route('/seed_stage_home')
def seed_stage_home():
    return render_template('seed_stage_home.html')

@app.route('/seed_stage_more_info',methods=["POST"])
def seed_stage_more_info():
    num_trains = int(request.form.get('num_trains'))
    return render_template('seed_stage_more_info.html',num_trains=num_trains)

@app.route('/new_seed_stage',methods=["POST"])
def new_seed_stage():
    processes = db.execute("SELECT process_version_name_used_as_template FROM pd_upstream.default_condition_process order by process_version_name_used_as_template")
    condition_versions = []
    for v in processes:
        if not v[0] in condition_versions:
            condition_versions.append(v[0])
    num_trains = int(request.form.get('num_trains'))
    seed_train_run_id_lst = []
    num_stages_lst = []
    num_timepoints_list = []
    for i in range(num_trains):
        seed_train_run_id = request.form.get('seed_train' + str(i))
        seed_train_run_id_lst.append(seed_train_run_id)
        num_stages = int(request.form.get('num_stages' + str(i)))
        num_stages_lst.append(num_stages)
        curr_lst = []
        for j in range(num_stages):
            curr_num_timepoints = int(request.form.get('num_timepoints_' + str(i) + '_' + str(j)))
            curr_lst.append(curr_num_timepoints)
        num_timepoints_list.append(curr_lst)
    max_num_stages = max(num_stages_lst)
    data = db.execute("SELECT * FROM pd_upstream.default_condition_process")
    rows = [dict(row) for row in data.fetchall()]
    form3 = SeedGrowthData()
    return render_template('new_seed_stage.html',form3=form3,rows=rows,condition_versions=condition_versions,max_num_stages=max_num_stages,num_timepoints_list=num_timepoints_list,num_trains=num_trains,seed_train_run_id_lst=seed_train_run_id_lst,num_stages_lst=num_stages_lst)

@app.route('/seed_stage_submitted',methods=["POST"])
def seed_stage_submitted():
    num_trains = int(request.form.get('num_trains')) # fix this
    # push cell bank data
    for i in range(num_trains):
        seed_train_run_id = request.form.get("4-" + str(1+0*num_trains + i))
        strain_number = request.form.get("4-" + str(1+1*num_trains + i))
        cell_bank_type = request.form.get("4-" + str(1+2*num_trains + i))
        cell_bank_lot_number = request.form.get("4-" + str(1+3*num_trains + i))
        cell_bank_vial_number = request.form.get("4-" + str(1+4*num_trains + i))
        date_cell_bank_made = request.form.get("4-" + str(1+5*num_trains + i))
        if date_cell_bank_made == "":
            date_cell_bank_made = None
        db.execute("INSERT INTO pd_upstream.cell_bank (seed_train_run_id,strain_number,cell_bank_type,cell_bank_lot_number,cell_bank_vial_number,date_cell_bank_made) \
                VALUES (:seed_train_run_id,:strain_number,:cell_bank_type,:cell_bank_lot_number,:cell_bank_vial_number,:date_cell_bank_made)",
        {"seed_train_run_id":seed_train_run_id,"strain_number":strain_number,"cell_bank_type":cell_bank_type,"cell_bank_lot_number":cell_bank_lot_number,"cell_bank_vial_number":cell_bank_vial_number,"date_cell_bank_made":date_cell_bank_made}) 
    db.commit()
    # push seed process condition data
    last_id_q = db.execute("SELECT id FROM pd_upstream.cell_bank ORDER BY id DESC LIMIT {}".format(num_trains))
    last_ids = []
    for i in last_id_q:
        last_ids.append(int(i[0]))
    last_ids.reverse()
    num_stages_lst = []
    seed_train_run_id_lst = []
    process_version_name_used_as_template_lst = []
    process_version_name_lst = []
    for i in range(num_trains):
        num_stages_lst.append(int(request.form.get('num_stages' + str(i))))
        seed_train_run_id_lst.append(request.form.get('5-' + str(i + 1)))
        process_version_name_used_as_template_lst.append(request.form.get('5-' + str(num_trains + i + 1)))
        process_version_name_lst.append(request.form.get('5-' + str(3*num_trains + i + 1)))
    for j in range(num_trains):
        for k in range(num_stages_lst[j]):
            starting_index = 4*num_trains + 1 + j + 16*num_trains*k
            seed_train_run_id = seed_train_run_id_lst[j]
            process_version_name_used_as_template = process_version_name_used_as_template_lst[j]
            process_version_name = process_version_name_lst[j]
            seed_stage_run_id = request.form.get('5-' + str(starting_index + 0*num_trains))
            seed_stage = request.form.get('5-' + str(starting_index + 1*num_trains))
            strain_number = request.form.get('5-' + str(starting_index + 2*num_trains))
            seed_stage_run_id_used_to_inoculate = request.form.get('5-' + str(starting_index + 3*num_trains))
            media = request.form.get('5-' + str(starting_index + 4*num_trains))
            media_lot_number = request.form.get('5-' + str(starting_index + 5*num_trains))
            temperature_c_setpoint = request.form.get('5-' + str(starting_index + 6*num_trains))
            if temperature_c_setpoint == "":
                temperature_c_setpoint = None
            agitation_setpoint_rpm = request.form.get('5-' + str(starting_index + 7*num_trains))
            if agitation_setpoint_rpm == "":
                agitation_setpoint_rpm = None
            orbital_diameter_cm = request.form.get('5-' + str(starting_index + 8*num_trains))
            if orbital_diameter_cm == "":
                orbital_diameter_cm = None
            do_percent_setpoint = request.form.get('5-' + str(starting_index + 9*num_trains))
            if do_percent_setpoint == "":
                do_percent_setpoint = None
            vessel = request.form.get('5-' + str(starting_index + 10*num_trains))
            vessel_size_ml = request.form.get('5-' + str(starting_index + 11*num_trains))
            if vessel_size_ml == "":
                vessel_size_ml = None
            target_seeding_density_percent_vv = request.form.get('5-' + str(starting_index + 12*num_trains))
            if target_seeding_density_percent_vv == "":
                target_seeding_density_percent_vv = None
            total_batch_volume_ml = request.form.get('5-' + str(starting_index + 13*num_trains))
            if total_batch_volume_ml == "":
                total_batch_volume_ml = None
            amount_of_sf_media_to_add_ml = request.form.get('5-' + str(starting_index + 14*num_trains))
            if amount_of_sf_media_to_add_ml == "":
                amount_of_sf_media_to_add_ml = None
            inoculum_size_to_add_ml = request.form.get('5-' + str(starting_index + 15*num_trains))
            if inoculum_size_to_add_ml == "":
                inoculum_size_to_add_ml = None
            db.execute("INSERT INTO pd_upstream.seed_process_condition (seed_train_run_id,process_version_name_used_as_template,process_version_name,seed_stage_run_id,seed_stage,strain_number,seed_stage_run_id_used_to_inoculate,media,media_lot_number,temperature_c_setpoint,agitation_setpoint_rpm,orbital_diameter_cm,do_percent_setpoint,vessel,vessel_size_ml,target_seeding_density_percent_vv,total_batch_volume_ml,amount_of_sf_media_to_add_ml,inoculum_size_to_add_ml,actual_seed_train_id) \
                VALUES (:seed_train_run_id,:process_version_name_used_as_template,:process_version_name,:seed_stage_run_id,:seed_stage,:strain_number,:seed_stage_run_id_used_to_inoculate,:media,:media_lot_number,:temperature_c_setpoint,:agitation_setpoint_rpm,:orbital_diameter_cm,:do_percent_setpoint,:vessel,:vessel_size_ml,:target_seeding_density_percent_vv,:total_batch_volume_ml,:amount_of_sf_media_to_add_ml,:inoculum_size_to_add_ml,:last_id)",
            {"seed_train_run_id":seed_train_run_id,"process_version_name_used_as_template":process_version_name_used_as_template,"process_version_name":process_version_name,"seed_stage_run_id":seed_stage_run_id,"seed_stage":seed_stage,"strain_number":strain_number,"seed_stage_run_id_used_to_inoculate":seed_stage_run_id_used_to_inoculate,"media":media,"media_lot_number":media_lot_number,"temperature_c_setpoint":temperature_c_setpoint,"agitation_setpoint_rpm":agitation_setpoint_rpm,"orbital_diameter_cm":orbital_diameter_cm,"do_percent_setpoint":do_percent_setpoint,"vessel":vessel,"vessel_size_ml":vessel_size_ml,"target_seeding_density_percent_vv":target_seeding_density_percent_vv,"total_batch_volume_ml":total_batch_volume_ml,"amount_of_sf_media_to_add_ml":amount_of_sf_media_to_add_ml,"inoculum_size_to_add_ml":inoculum_size_to_add_ml,"last_id":last_ids[j]}) 
            db.commit()
    # push seed growth data
    tot_num_stages = sum(num_stages_lst)
    last_id_q = db.execute("SELECT id FROM pd_upstream.seed_process_condition ORDER BY id DESC LIMIT {}".format(tot_num_stages))
    last_ids = []
    for i in last_id_q:
        last_ids.append(int(i[0]))
    last_ids.reverse()
    num_additions_list = []
    for i in range(num_trains):
        curr_lst = []
        for j in range(num_stages_lst[i]):
            curr_lst.append(int(request.form.get('num_timepoints_' + str(i) + '_' + str(j))))
        num_additions_list.append(curr_lst)
    first_elem = 1
    curr_stage = 0
    curr_row = 0
    for i in range(num_trains):
        for j in range(num_stages_lst[i]):
            for k in range(num_additions_list[i][j]):
                seed_train_run_id = request.form.get('6-' + str(first_elem + 0))
                seed_stage_run_id = request.form.get('6-' + str(first_elem + 1))
                ferm_sample_id = request.form.get('6-' + str(first_elem + 2))
                seed_stage = request.form.get('6-' + str(first_elem + 3))
                if seed_stage == "":
                    seed_stage = None
                description = request.form.get('6-' + str(first_elem + 4))
                datetime = request.form.get('6-' + str(first_elem + 5))
                if datetime == "":
                    datetime = None
                take_in_or_out_of_incubator = request.form.get('6-select' + str(curr_row))
                temperature_seed_stage_held_at_c = request.form.get('6-' + str(first_elem + 6))
                if temperature_seed_stage_held_at_c == "":
                    temperature_seed_stage_held_at_c = None
                total_elapsed_time_h = request.form.get('6-' + str(first_elem + 7))
                if total_elapsed_time_h == "":
                    total_elapsed_time_h = None
                true_eft_h = request.form.get('6-' + str(first_elem + 8))
                if true_eft_h == "":
                    true_eft_h = None
                notes = request.form.get('6-' + str(first_elem + 9))
                ysi_glucose_gl = request.form.get('6-' + str(first_elem + 10))
                if ysi_glucose_gl == "":
                    ysi_glucose_gl = None
                o2_flow_percent = request.form.get('6-' + str(first_elem + 11))
                if o2_flow_percent == "":
                    o2_flow_percent = None
                total_gas_flow_slpm = request.form.get('6-' + str(first_elem + 12))
                if total_gas_flow_slpm == "":
                    total_gas_flow_slpm = None
                base_totalizer_ml = request.form.get('6-' + str(first_elem + 13))
                if base_totalizer_ml == "":
                    base_totalizer_ml = None
                online_ph = request.form.get('6-' + str(first_elem + 14))
                if online_ph == "":
                    online_ph = None
                offline_ph = request.form.get('6-' + str(first_elem + 15))
                if offline_ph == "":
                    offline_ph = None
                od595 = request.form.get('6-' + str(first_elem + 16))
                if od595 == "":
                    od595 = None
                specific_growth_rate_k_1 = request.form.get('6-' + str(first_elem + 17))
                if specific_growth_rate_k_1 == "":
                    specific_growth_rate_k_1 = None
                od_bioht = request.form.get('6-' + str(first_elem + 18))
                if od_bioht == "":
                    od_bioht = None
                bioht_glucose_g_l = request.form.get('6-' + str(first_elem + 19))
                if bioht_glucose_g_l == "":
                    bioht_glucose_g_l = None
                bioht_lactate_mmol = request.form.get('6-' + str(first_elem + 20))
                if bioht_lactate_mmol == "":
                    bioht_lactate_mmol = None
                bioht_acetate_mmol = request.form.get('6-' + str(first_elem + 21))
                if bioht_acetate_mmol == "":
                    bioht_acetate_mmol = None
                bioht_ammonia_mmol = request.form.get('6-' + str(first_elem + 22))
                if bioht_ammonia_mmol == "":
                    bioht_ammonia_mmol = None
                bioht_glutamine_mmol = request.form.get('6-' + str(first_elem + 23))
                if bioht_glutamine_mmol == "":
                    bioht_glutamine_mmol = None
                bioht_glutamate_mmol = request.form.get('6-' + str(first_elem + 24))
                if bioht_glutamate_mmol == "":
                    bioht_glutamate_mmol = None
                bioht_phosphate_mmol = request.form.get('6-' + str(first_elem + 25))
                if bioht_phosphate_mmol == "":
                    bioht_phosphate_mmol = None
                bioht_magnesium_mmol = request.form.get('6-' + str(first_elem + 26))
                if bioht_magnesium_mmol == "":
                    bioht_magnesium_mmol = None
                bioht_formate_mg_l = request.form.get('6-' + str(first_elem + 27))
                if bioht_formate_mg_l == "":
                    bioht_formate_mg_l = None
                bioht_pyruvate_mmol = request.form.get('6-' + str(first_elem + 28))
                if bioht_pyruvate_mmol == "":
                    bioht_pyruvate_mmol = None
                bioht_ldh_u_l = request.form.get('6-' + str(first_elem + 29))
                if bioht_ldh_u_l == "":
                    bioht_ldh_u_l = None
                bioht_igg_mg_l = request.form.get('6-' + str(first_elem + 30))
                if bioht_igg_mg_l == "":
                    bioht_igg_mg_l = None
                bioht_total_protein_g_l = request.form.get('6-' + str(first_elem + 31))
                if bioht_total_protein_g_l == "":
                    bioht_total_protein_g_l = None
                sodium = request.form.get('6-' + str(first_elem + 32))
                if sodium == "":
                    sodium = None
                potassium = request.form.get('6-' + str(first_elem + 33))
                if potassium == "":
                    potassium = None
                osm_mosm_kg = request.form.get('6-' + str(first_elem + 34))
                if osm_mosm_kg == "":
                    osm_mosm_kg = None
                first_elem += 35
                db.execute("INSERT INTO pd_upstream.seed_growth_data (seed_train_run_id,seed_stage_run_id,ferm_sample_id,seed_stage,description,datetime,take_in_or_out_of_incubator,temperature_seed_stage_held_at_c,total_elapsed_time_h,true_eft_h,notes,ysi_glucose_gl,o2_flow_percent,total_gas_flow_slpm,base_totalizer_ml,online_ph,offline_ph,od595,specific_growth_rate_k_1,od_bioht,bioht_glucose_g_l,bioht_lactate_mmol,bioht_acetate_mmol,bioht_ammonia_mmol,bioht_glutamine_mmol,bioht_glutamate_mmol,bioht_phosphate_mmol,bioht_magnesium_mmol,bioht_formate_mg_l,bioht_pyruvate_mmol,bioht_ldh_u_l,bioht_igg_mg_l,bioht_total_protein_g_l,sodium,potassium,osm_mosm_kg,actual_seed_stage_id) \
                            VALUES (:seed_train_run_id,:seed_stage_run_id,:ferm_sample_id,:seed_stage,:description,:datetime,:take_in_or_out_of_incubator,:temperature_seed_stage_held_at_c,:total_elapsed_time_h,:true_eft_h,:notes,:ysi_glucose_gl,:o2_flow_percent,:total_gas_flow_slpm,:base_totalizer_ml,:online_ph,:offline_ph,:od595,:specific_growth_rate_k_1,:od_bioht,:bioht_glucose_g_l,:bioht_lactate_mmol,:bioht_acetate_mmol,:bioht_ammonia_mmol,:bioht_glutamine_mmol,:bioht_glutamate_mmol,:bioht_phosphate_mmol,:bioht_magnesium_mmol,:bioht_formate_mg_l,:bioht_pyruvate_mmol,:bioht_ldh_u_l,:bioht_igg_mg_l,:bioht_total_protein_g_l,:sodium,:potassium,:osm_mosm_kg,:last_id)",
                {"seed_train_run_id":seed_train_run_id,"seed_stage_run_id":seed_stage_run_id,"ferm_sample_id":ferm_sample_id,"seed_stage":seed_stage,"description":description,"datetime":datetime,"take_in_or_out_of_incubator":take_in_or_out_of_incubator,"temperature_seed_stage_held_at_c":temperature_seed_stage_held_at_c,"total_elapsed_time_h":total_elapsed_time_h,"true_eft_h":true_eft_h,"notes":notes,"ysi_glucose_gl":ysi_glucose_gl,"o2_flow_percent":o2_flow_percent,"total_gas_flow_slpm":total_gas_flow_slpm,"base_totalizer_ml":base_totalizer_ml,"online_ph":online_ph,"offline_ph":offline_ph,"od595":od595,"specific_growth_rate_k_1":specific_growth_rate_k_1,"od_bioht":od_bioht,"bioht_glucose_g_l":bioht_glucose_g_l,"bioht_lactate_mmol":bioht_lactate_mmol,"bioht_acetate_mmol":bioht_acetate_mmol,"bioht_ammonia_mmol":bioht_ammonia_mmol,"bioht_glutamine_mmol":bioht_glutamine_mmol,"bioht_glutamate_mmol":bioht_glutamate_mmol,"bioht_phosphate_mmol":bioht_phosphate_mmol,"bioht_magnesium_mmol":bioht_magnesium_mmol,"bioht_formate_mg_l":bioht_formate_mg_l,"bioht_pyruvate_mmol":bioht_pyruvate_mmol,"bioht_ldh_u_l":bioht_ldh_u_l,"bioht_igg_mg_l":bioht_igg_mg_l,"bioht_total_protein_g_l":bioht_total_protein_g_l,"sodium":sodium,"potassium":potassium,"osm_mosm_kg":osm_mosm_kg,"last_id":last_ids[curr_stage]}) 
                db.commit()
                curr_row += 1
            curr_stage += 1
    return render_template('submitted.html')

@app.route('/condition_process_add')
def condition_process_add():
    return render_template('condition_process_add.html')

@app.route('/condition_process_add_submitted',methods=["POST"])
def condition_process_add_submitted():
    process_version_name_used_as_template = request.form.get('process_version_name_used_as_template')
    seed_stage_run_id = request.form.get('seed_stage_run_id')
    seed_stage = request.form.get('seed_stage')
    media = request.form.get('media')
    temperature_c_setpoint = request.form.get('temperature_c_setpoint')
    if temperature_c_setpoint == "":
        temperature_c_setpoint = None
    agitation_setpoint_rpm = request.form.get('agitation_setpoint_rpm')
    if agitation_setpoint_rpm == "":
        agitation_setpoint_rpm = None
    orbital_diameter_cm = request.form.get('orbital_diameter_cm')
    if orbital_diameter_cm == "":
        orbital_diameter_cm = None
    do_percent_setpoint = request.form.get('do_percent_setpoint')
    if do_percent_setpoint == "":
        do_percent_setpoint = None
    target_seeding_density_percent_vv = request.form.get('target_seeding_density_percent_vv')
    if target_seeding_density_percent_vv == "":
        target_seeding_density_percent_vv = None
    db.execute("INSERT INTO pd_upstream.default_condition_process (seed_stage_run_id,seed_stage,temperature_c_setpoint,process_version_name_used_as_template,media,agitation_setpoint_rpm,orbital_diameter_cm,do_percent_setpoint,target_seeding_density_percent_vv) \
                VALUES (:seed_stage_run_id,:seed_stage,:temperature_c_setpoint,:process_version_name_used_as_template,:media,:agitation_setpoint_rpm,:orbital_diameter_cm,:do_percent_setpoint,:target_seeding_density_percent_vv)",
        {"seed_stage_run_id":seed_stage_run_id,"seed_stage":seed_stage,"temperature_c_setpoint":temperature_c_setpoint,"process_version_name_used_as_template":process_version_name_used_as_template,"media":media,"agitation_setpoint_rpm":agitation_setpoint_rpm,"orbital_diameter_cm":orbital_diameter_cm,"do_percent_setpoint":do_percent_setpoint,"target_seeding_density_percent_vv":target_seeding_density_percent_vv}) 
    conn.commit()
    return render_template('submitted.html')

@app.route('/condition_process_edit')
def condition_process_edit():
    processes = db.execute("SELECT process_version_name_used_as_template FROM pd_upstream.default_condition_process order by process_version_name_used_as_template")
    condition_versions = []
    for v in processes:
        if not v[0] in condition_versions:
            condition_versions.append(v[0])
    data = db.execute("SELECT * FROM pd_upstream.default_condition_process")
    rows = [dict(row) for row in data.fetchall()]
    return render_template('condition_process_edit.html',rows=rows,condition_versions=condition_versions)

@app.route('/condition_process_edit_submitted',methods=["POST"])
def condition_process_edit_submitted():
    process_version_name_used_as_template = request.form.get('process_version_name_used_as_template')
    seed_stage_run_id = request.form.get('seed_stage_run_id')
    seed_stage = request.form.get('seed_stage')
    media = request.form.get('media')
    temperature_c_setpoint = request.form.get('temperature_c_setpoint')
    if temperature_c_setpoint == "":
        temperature_c_setpoint = None
    agitation_setpoint_rpm = request.form.get('agitation_setpoint_rpm')
    if agitation_setpoint_rpm == "":
        agitation_setpoint_rpm = None
    orbital_diameter_cm = request.form.get('orbital_diameter_cm')
    if orbital_diameter_cm == "":
        orbital_diameter_cm = None
    do_percent_setpoint = request.form.get('do_percent_setpoint')
    if do_percent_setpoint == "":
        do_percent_setpoint = None
    target_seeding_density_percent_vv = request.form.get('target_seeding_density_percent_vv')
    if target_seeding_density_percent_vv == "":
        target_seeding_density_percent_vv = None
    s = "DELETE FROM pd_upstream.default_condition_process WHERE process_version_name_used_as_template='{}';".format(process_version_name_used_as_template)
    db.execute(s)
    db.execute("INSERT INTO pd_upstream.default_condition_process (seed_stage_run_id,seed_stage,temperature_c_setpoint,process_version_name_used_as_template,media,agitation_setpoint_rpm,orbital_diameter_cm,do_percent_setpoint,target_seeding_density_percent_vv) \
                VALUES (:seed_stage_run_id,:seed_stage,:temperature_c_setpoint,:process_version_name_used_as_template,:media,:agitation_setpoint_rpm,:orbital_diameter_cm,:do_percent_setpoint,:target_seeding_density_percent_vv)",
        {"seed_stage_run_id":seed_stage_run_id,"seed_stage":seed_stage,"temperature_c_setpoint":temperature_c_setpoint,"process_version_name_used_as_template":process_version_name_used_as_template,"media":media,"agitation_setpoint_rpm":agitation_setpoint_rpm,"orbital_diameter_cm":orbital_diameter_cm,"do_percent_setpoint":do_percent_setpoint,"target_seeding_density_percent_vv":target_seeding_density_percent_vv}) 
    db.commit()
    return render_template('submitted.html')
